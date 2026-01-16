"""
Enhanced Test Case Capture Application
Automatically captures test cases with browser URL monitoring and module detection
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import threading
import time
import sys
import re
import json
import subprocess
import os
import tempfile
from urllib.parse import urlparse, parse_qs

try:
    from pynput import mouse, keyboard
    from pynput.mouse import Listener as MouseListener
    from pynput.keyboard import Listener as KeyboardListener, Key
    PYNPUT_AVAILABLE = True
except ImportError:
    PYNPUT_AVAILABLE = False
    Key = None

try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

# Base URL to filter - will be set dynamically from first detected URL
BASE_URL = None  # Will be set when first URL is detected

# Module identification patterns
MODULE_PATTERNS = {
    "login": [r"/login", r"login", r"signin", r"sign-in"],
    "advertiser dashboard": [r"/advertiser", r"/dashboard", r"advertiser.*dashboard"],
    "brand dashboard": [r"/brand", r"brand.*dashboard"],
    "manage payments": [r"/payment", r"/payments", r"payment"],
    "manage users": [r"/user", r"/users", r"user.*management"],
    "manage accounts": [r"/account", r"/accounts", r"account.*management"],
    "target": [r"/target", r"targeting"],
    "plan": [r"/plan", r"/planning", r"plan"],
    "activate": [r"/activate", r"activation"],
    "measure": [r"/measure", r"/measurement", r"measure", r"/analytics", r"/report"]
}


class BrowserMonitor:
    """Monitors browser URLs and navigation"""
    
    def __init__(self, callback, base_url=BASE_URL):
        self.callback = callback
        self.base_url = base_url
        self.monitoring = False
        self.current_url = ""
        self.last_url = ""
        self.last_valid_url = None  # Track last valid URL for action capture continuity
        self.current_module = ""
        self.current_page = ""
        self.driver = None
        self.monitor_thread = None
        self.current_tab = None
        
    def start_monitoring(self):
        """Start monitoring browser URLs"""
        self.monitoring = True
        self.monitor_thread = threading.Thread(target=self._monitor_urls, daemon=True)
        self.monitor_thread.start()
        return True
    
    def stop_monitoring(self):
        """Stop monitoring browser URLs"""
        self.monitoring = False
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
            self.driver = None
    
    def _monitor_urls(self):
        """Monitor URLs from active browser tabs"""
        last_check_time = 0
        check_count = 0
        last_window_title = ""  # Track window title for tab switch detection
        while self.monitoring:
            try:
                current_time = time.time()
                # Check URL every 2 seconds
                if current_time - last_check_time >= 2:
                    last_check_time = current_time
                    check_count += 1
                    
                    # Log periodic check (every 10 checks = 20 seconds)
                    if check_count % 10 == 0:
                        if hasattr(self, 'callback') and self.callback:
                            # Try to log through callback if available
                            pass
                    
                    # Primary method: Chrome DevTools Protocol
                    url = self._get_url_from_chrome_devtools()
                    if url:
                        # Always update if URL is different (even slightly) - regardless of domain
                        if url != self.current_url:
                            self._handle_url_change(url)
                        # Also update if current_url is empty
                        elif not self.current_url:
                                self._handle_url_change(url)
                    elif not url and self.current_url:
                        # URL might have changed to non-target URL, but keep monitoring
                        pass
                    
                    # Fallback: Try to extract from window title (less reliable)
                    if not url and sys.platform == "win32":
                        url = self._get_url_from_window_title()
                        if url and url != self.current_url:
                            # Allow URL changes from window title regardless of base_url
                            self._handle_url_change(url)
                    
                    # Also monitor window title for tab switches (when URL doesn't change)
                    if sys.platform == "win32" and url and url.startswith(self.base_url):
                        try:
                            import win32gui
                            hwnd = win32gui.GetForegroundWindow()
                            current_title = win32gui.GetWindowText(hwnd)
                            
                            # Check if title changed (might indicate tab switch)
                            if current_title and current_title != last_window_title and last_window_title:
                                # Title changed - check if it's a tab switch
                                tab_keywords = ["Accounts", "Users", "Notification"]
                                for keyword in tab_keywords:
                                    if keyword in current_title and keyword not in last_window_title:
                                        # Tab switch detected!
                                        if self.callback:
                                            navigation_action = f"Tab switched to '{keyword}' (URL unchanged: {url})"
                                            self.callback(navigation_action, url, self.current_module, self.current_page)
                                        break
                            
                            last_window_title = current_title
                        except:
                            pass
                        
            except Exception as e:
                pass
            
            time.sleep(1)  # Check every second
    
    def _get_url_from_window_title(self):
        """Try to extract URL from browser window title (fallback method) - improved to check all browser windows"""
        try:
            if sys.platform == "win32":
                import win32gui
                import re
                
                # Get all visible windows
                windows = []
                def enum_windows_callback(hwnd, param):
                    """Callback to enumerate all windows"""
                    try:
                        if win32gui.IsWindowVisible(hwnd):
                            window_title = win32gui.GetWindowText(hwnd)
                            if window_title:
                                param.append((hwnd, window_title))
                    except:
                        pass
                    return True
                
                win32gui.EnumWindows(enum_windows_callback, windows)
                
                # Look for browser windows (Chrome, Edge, Firefox, etc.)
                browser_keywords = ['chrome', 'edge', 'firefox', 'opera', 'brave', 'vivaldi', 'microsoft edge']
                url_pattern = r'https?://[^\s<>"\'\)]+'
                
                browser_windows_found = []
                # First, collect all browser windows
                for hwnd, window_title in windows:
                    if any(keyword in window_title.lower() for keyword in browser_keywords):
                        browser_windows_found.append((hwnd, window_title))
                
                # Try to find URL in any browser window title
                for hwnd, window_title in browser_windows_found:
                        # Try to extract URL from title
                        match = re.search(url_pattern, window_title)
                        if match:
                            url = match.group(0).rstrip('.,;:!?)')
                            # Return any valid URL (generic - not restricted to base_url)
                            if url.startswith('http://') or url.startswith('https://'):
                                # Check if it's a valid URL format
                                domain_part = url.split('://')[1].split('/')[0]
                                if '.' in domain_part and len(domain_part) > 3:  # Has valid domain
                                    return url
                
                # Also check foreground window specifically (even if not in browser_windows_found)
                hwnd = win32gui.GetForegroundWindow()
                window_title = win32gui.GetWindowText(hwnd)
                if window_title:
                    # Check if it's a browser window
                    if any(keyword in window_title.lower() for keyword in browser_keywords):
                        match = re.search(url_pattern, window_title)
                        if match:
                            url = match.group(0).rstrip('.,;:!?)')
                            if url.startswith('http://') or url.startswith('https://'):
                                domain_part = url.split('://')[1].split('/')[0]
                                if '.' in domain_part and len(domain_part) > 3:
                                    return url
        except Exception as e:
            # Log error for debugging
            if hasattr(self, 'log_message'):
                self.log_message(f"Window title detection error: {e}", "WARNING")
        return None
    
    def _get_url_from_chrome_devtools(self):
        """Get current URL using Chrome DevTools Protocol - returns exact active tab URL"""
        try:
            import json as json_lib
            import urllib.request
            
            # Try multiple common ports
            ports = [9222, 9223, 9224, 9225, 9226]  # Added more ports
            for port in ports:
                try:
                    response = urllib.request.urlopen(f"http://localhost:{port}/json", timeout=1)  # Faster timeout
                    tabs = json_lib.loads(response.read().decode())
                    
                    if not tabs:
                        continue
                    
                    # Strategy 1: Find the active/focused tab (most reliable)
                    # Get the currently active window title to match with tab
                    active_window_title = None
                    if sys.platform == "win32":
                        try:
                            import win32gui
                            hwnd = win32gui.GetForegroundWindow()
                            active_window_title = win32gui.GetWindowText(hwnd)
                        except:
                            pass
                    
                    active_tab = None
                    active_tab_title_match = None
                    
                    for tab in tabs:
                        url = tab.get('url', '')
                        tab_type = tab.get('type', '')
                        title = tab.get('title', '')
                        
                        # Check if this is a page tab with a valid URL
                        if tab_type == 'page' and url and (url.startswith('http://') or url.startswith('https://')):
                            # If we have active window title, try to match it with tab title
                            if active_window_title and title:
                                if title.lower() in active_window_title.lower() or active_window_title.lower() in title.lower():
                                    # This tab matches the active window - use it!
                                    return url
                            
                            # Prefer tabs that have webSocketDebuggerUrl (usually means they're active)
                            if 'webSocketDebuggerUrl' in tab and tab['webSocketDebuggerUrl']:
                                if not active_tab:
                                    active_tab = url
                                    active_tab_title_match = title
                            
                            # If no active tab found yet, use the first valid page tab
                            if not active_tab:
                                active_tab = url
                                active_tab_title_match = title
                    
                    if active_tab:
                        return active_tab
                    
                    # Strategy 2: If base_url is set, prioritize tabs matching base_url
                    if self.base_url:
                        matching_base_url = []
                    for tab in tabs:
                        url = tab.get('url', '')
                        tab_type = tab.get('type', '')
                        if tab_type == 'page' and url and url.startswith(self.base_url):
                            matching_base_url.append(url)
                        
                        if matching_base_url:
                            return matching_base_url[0]  # Return first match
                    
                    # Strategy 3: Return first valid page tab URL (fallback - return ANY valid URL)
                    # This ensures we detect the URL even if domain doesn't match
                    for tab in tabs:
                        url = tab.get('url', '')
                        tab_type = tab.get('type', '')
                        # Return any page tab with a valid URL
                        if tab_type == 'page' and url:
                            # Return any HTTP/HTTPS URL (be lenient)
                            if url.startswith('http://') or url.startswith('https://'):
                                # Basic validation - just check it has a domain
                                try:
                                    url_parts = url.split('://')
                                    if len(url_parts) == 2:
                                        domain_part = url_parts[1].split('/')[0].split(':')[0]  # Remove port if present
                                        # Valid domain should have at least one dot and be longer than 3 chars
                                        if '.' in domain_part and len(domain_part) > 3:
                                            return url  # Return first valid URL found
                                except:
                                    # If parsing fails, still return it if it starts with http/https
                                    return url
                            
                except urllib.error.URLError:
                    # Port not available - this is normal for manually opened browsers
                    continue
                except Exception as e:
                    continue
        except:
            pass
        return None
    
    def _get_url_from_selenium(self):
        """Get URL using Selenium (if browser is controlled by Selenium)"""
        # This would work if the browser was opened by Selenium
        # For monitoring existing browsers, we need Chrome DevTools Protocol
        return None
    
    def _handle_url_change(self, url):
        """Handle URL change event"""
        # Prevent processing the same URL change multiple times
        if url == self.current_url:
            return  # URL hasn't actually changed, skip processing
        
        from urllib.parse import urlparse
        global BASE_URL
        
        # Extract domain from the new URL
        parsed = urlparse(url)
        new_base_url = f"{parsed.scheme}://{parsed.netloc}"
        
        # If base_url is not set, set it from the first URL detected
        if not self.base_url:
            self.base_url = new_base_url
            # Update global BASE_URL for consistency
            BASE_URL = self.base_url
        # If URL is from a different domain, update base_url to allow domain changes
        elif new_base_url != self.base_url:
            # Allow domain changes - update base_url to the new domain
            # This handles cases where user navigates to a different site
            self.base_url = new_base_url
            BASE_URL = self.base_url
        
        # Always process URL changes (removed the restriction that blocked different domains)
        self.last_url = self.current_url
        self.current_url = url
        # Track last valid URL to help with action capture after login
        self.last_valid_url = url
        
        # Identify module and page from URL
        module, page = self._identify_module_and_page(url)
        self.current_module = module
        self.current_page = page
        
        # Notify callback
        if self.callback:
            navigation_action = f"Navigated to: {url}\nModule: {module} | Page: {page}"
            self.callback(navigation_action, url, module, page)
    
    def _identify_module_and_page(self, url):
        """Identify module and page from URL"""
        url_lower = url.lower()
        path = urlparse(url).path.lower()
        
        # Identify module
        module = "General"
        for mod_name, patterns in MODULE_PATTERNS.items():
            for pattern in patterns:
                if re.search(pattern, path) or re.search(pattern, url_lower):
                    module = mod_name.title()
                    break
            if module != "General":
                break
        
        # Identify page from URL path
        page = "Home"
        path_parts = [p for p in path.split('/') if p]
        if path_parts:
            # Use last meaningful path segment as page name
            page = path_parts[-1].replace('-', ' ').replace('_', ' ').title()
            if not page or page == '':
                page = path_parts[0].replace('-', ' ').replace('_', ' ').title() if path_parts else "Home"
        else:
            page = "Home"
        
        return module, page
    
    def get_current_info(self):
        """Get current URL, module, and page"""
        return {
            'url': self.current_url,
            'module': self.current_module,
            'page': self.current_page
        }


class ActionMonitor:
    """Monitors user actions and captures them automatically"""
    
    def __init__(self, callback, browser_monitor=None):
        self.callback = callback
        self.browser_monitor = browser_monitor
        self.monitoring = False
        self.mouse_listener = None
        self.keyboard_listener = None
        self.last_action_time = time.time()
        self.action_buffer = []
        self.current_window = ""
        self.last_window = ""
        self.click_count = 0
        self.key_count = 0
        self.is_target_application = False  # Track if we're on target application
        self.manual_url_set = False  # Track if URL was set manually
        self.log_callback = None  # Callback for logging
        self.last_window_title = ""  # Track window title to detect tab switches
        self.last_title_check_time = 0  # Throttle title checks
        
    def start_monitoring(self):
        """Start monitoring actions"""
        if not PYNPUT_AVAILABLE:
            return False
        
        self.monitoring = True
        self.last_action_time = time.time()
        
        # Initialize window title tracking
        self.last_window_title = self._get_window_title()
        
        # Start mouse listener
        self.mouse_listener = MouseListener(
            on_click=self.on_mouse_click,
            on_scroll=self.on_scroll
        )
        self.mouse_listener.start()
        
        # Start keyboard listener
        self.keyboard_listener = KeyboardListener(
            on_press=self.on_key_press,
            on_release=self.on_key_release
        )
        self.keyboard_listener.start()
        
        # Start window monitoring thread
        if PSUTIL_AVAILABLE:
            window_thread = threading.Thread(target=self.monitor_windows, daemon=True)
            window_thread.start()
        
        return True
    
    def stop_monitoring(self):
        """Stop monitoring actions"""
        self.monitoring = False
        if self.mouse_listener:
            self.mouse_listener.stop()
        if self.keyboard_listener:
            self.keyboard_listener.stop()
    
    def _check_target_application(self):
        """Check if current window is target application"""
        # If URL was manually set, trust the user and always return True
        if self.manual_url_set:
            return True
        
        # Check if current URL matches base URL
        if self.browser_monitor and self.browser_monitor.current_url:
            url = self.browser_monitor.current_url
            base_url = self.browser_monitor.base_url or BASE_URL
            # Allow if URL matches base URL (or if base_url not set yet, allow any URL)
            if not base_url or url.startswith(base_url):
                return True
            # Also allow if we're monitoring and URL was previously set (might be transitioning after login)
            if self.monitoring and hasattr(self.browser_monitor, 'last_valid_url'):
                if self.browser_monitor.last_valid_url:
                    if not base_url or self.browser_monitor.last_valid_url.startswith(base_url):
                        return True
        
        return False
    
    def on_mouse_click(self, x, y, button, pressed):
        """Handle mouse click events"""
        if not self.monitoring or not pressed:
            return
        
        # Check if we should capture this click
        is_target = self._check_target_application()
        has_url_set = self.browser_monitor and self.browser_monitor.current_url
        
        # Log click detection for debugging
        if hasattr(self, 'root') and self.root:
            if has_url_set:
                if is_target:
                    # Log successful capture
                    pass  # Will be logged in capture_action
                else:
                    # Log filtered click
                    url = self.browser_monitor.current_url
                    self.root.after(0, lambda: self._log_filtered_action(
                        f"Mouse click detected at ({x}, {y}) but filtered - URL '{url}' doesn't match base URL"))
            else:
                # Log that URL is not set
                self.root.after(0, lambda: self._log_filtered_action(
                    f"Mouse click detected at ({x}, {y}) but filtered - URL not set"))
        
        # Capture if on target OR if URL is manually set (trust user)
        # If URL was manually set, we trust the user and capture all clicks
        # Also capture if monitoring is active (even if URL check fails temporarily after login)
        should_capture = is_target or (has_url_set and self.manual_url_set) or (self.monitoring and has_url_set)
        
        if should_capture:
            self.click_count += 1
            # Capture click - could be button, dropdown, menu, link, etc.
            action = f"Mouse {button.name} click at ({x}, {y})"
            
            # Log click with more detail - identify potential dropdown/menu clicks
            if self.log_callback and self.root:
                # Log every click with context
                self.root.after(0, lambda: self.log_callback(
                    f"🖱️ Click captured at ({x}, {y}) - Could be button/dropdown/menu/link", "ACTION"))
            
            self.capture_action(action)
            
            # Check for tab switch after click (delay to allow title to update)
            if self.root:
                self.root.after(500, self._check_tab_switch_after_click)  # Check after 500ms
        else:
            # Log why click wasn't captured (only log occasionally to avoid spam)
            if self.click_count % 10 == 0:  # Log every 10th filtered click
                if has_url_set:
                    url = self.browser_monitor.current_url
                    if self.log_callback and self.root:
                        self.root.after(0, lambda: self.log_callback(
                            f"Click filtered - URL '{url}' doesn't match {BASE_URL}. Set URL manually to capture all clicks.", "WARNING"))
                else:
                    if self.log_callback and self.root:
                        self.root.after(0, lambda: self.log_callback(
                            f"Click filtered - URL not set. Please set URL in Manual Override section!", "WARNING"))
    
    def _log_filtered_action(self, message):
        """Log filtered actions for debugging"""
        if hasattr(self, 'log_callback') and self.log_callback:
            try:
                self.log_callback(message, "WARNING")
            except:
                pass
    
    def on_scroll(self, x, y, dx, dy):
        """Handle mouse scroll events"""
        if not self.monitoring:
            return
        
        # Only capture if on target application
        if not self._check_target_application():
            return
        
        direction = "down" if dy < 0 else "up"
        action = f"Scroll {direction} at ({x}, {y})"
        self.capture_action(action)
    
    def on_key_press(self, key):
        """Handle key press events"""
        if not self.monitoring:
            return
        
        # Check if we should capture
        is_target = self._check_target_application()
        has_url_set = self.browser_monitor and self.browser_monitor.current_url
        should_capture = is_target or (has_url_set and self.manual_url_set)
        
        if not should_capture:
            return
        
        try:
            # Handle special keys (Enter, Tab, etc.)
            if Key:
                special_keys = {
                    Key.enter: "Enter key pressed",
                    Key.tab: "Tab key pressed",
                    Key.space: "Space key pressed",
                    Key.backspace: "Backspace key pressed",
                    Key.delete: "Delete key pressed",
                    Key.esc: "Escape key pressed"
                }
            else:
                special_keys = {}
            
            if key in special_keys:
                action = special_keys[key]
                self.capture_action(action)
                if self.log_callback and self.root:
                    self.root.after(0, lambda: self.log_callback(f"Special key: {action}", "ACTION"))
                return
            
            # Handle regular character keys
            if hasattr(key, 'char') and key.char:
                self.key_count += 1
                
                # Log every keystroke in activity log (but don't spam actions)
                if self.log_callback and self.root:
                    # Only log every 5th keystroke to avoid spam
                    if self.key_count % 5 == 0:
                        self.root.after(0, lambda: self.log_callback(
                            f"Text input detected ({self.key_count} characters typed)", "ACTION"))
                
                # Capture text input more frequently (every 5 keystrokes instead of 20)
                if self.key_count % 5 == 0:
                    action = f"Text input entered ({self.key_count} characters)"
                    self.capture_action(action)
                # Also capture on first keystroke to show typing started
                elif self.key_count == 1:
                    action = "Started typing text"
                    self.capture_action(action)
        except AttributeError:
            # Handle special keys that don't have char attribute
            try:
                key_name = str(key).replace('Key.', '')
                if key_name not in ['ctrl', 'alt', 'shift', 'cmd']:
                    if self.log_callback and self.root:
                        self.root.after(0, lambda: self.log_callback(
                            f"Key pressed: {key_name}", "ACTION"))
            except:
                pass
    
    def on_key_release(self, key):
        """Handle key release events"""
        # Capture when user stops typing (releases Enter or Tab)
        if not self.monitoring:
            return
        
        is_target = self._check_target_application()
        has_url_set = self.browser_monitor and self.browser_monitor.current_url
        should_capture = is_target or (has_url_set and self.manual_url_set)
        
        if should_capture:
            try:
                if Key:
                    if key == Key.enter:
                        if self.key_count > 0:
                            action = f"Finished entering text ({self.key_count} characters total)"
                            self.capture_action(action)
                            if self.log_callback and self.root:
                                self.root.after(0, lambda: self.log_callback(
                                    f"Text entry completed: {self.key_count} characters", "ACTION"))
                            self.key_count = 0  # Reset counter
                    elif key == Key.tab:
                        if self.key_count > 0:
                            action = f"Tabbed after entering text ({self.key_count} characters)"
                            self.capture_action(action)
                            if self.log_callback and self.root:
                                self.root.after(0, lambda: self.log_callback(
                                    f"Tabbed after text entry: {self.key_count} characters", "ACTION"))
                            self.key_count = 0  # Reset counter
            except:
                pass
    
    def monitor_windows(self):
        """Monitor active window changes"""
        if not PSUTIL_AVAILABLE:
            return
        
        while self.monitoring:
            try:
                # Get active window (Windows-specific)
                if sys.platform == "win32":
                    try:
                        import win32gui
                        hwnd = win32gui.GetForegroundWindow()
                        window_title = win32gui.GetWindowText(hwnd)
                        
                        if window_title and window_title != self.current_window:
                            self.last_window = self.current_window
                            self.current_window = window_title
                            
                            # Update last_window_title for tab switch detection
                            if window_title != self.last_window_title:
                                old_title = self.last_window_title
                                self.last_window_title = window_title
                                
                                # Check for tab switch (same window, different tab)
                                if old_title and any(browser in window_title.lower() for browser in ['chrome', 'edge', 'microsoft edge', 'firefox']):
                                    if self._check_target_application():
                                        # Check if this is a tab switch (title changed but likely same URL)
                                        self._check_tab_switch_in_title(window_title, old_title)
                            
                            # Only capture if it's a browser window (might contain our app)
                            if any(browser in window_title.lower() for browser in ['chrome', 'edge', 'microsoft edge', 'firefox']):
                                if self.last_window:  # Don't capture initial window
                                    action = f"Switched to window: {window_title}"
                                    # Only capture if on target application
                                    if self._check_target_application():
                                        self.capture_action(action)
                    except ImportError:
                        pass  # win32gui not available
            except:
                pass
            
            time.sleep(0.5)  # Check every 500ms
    
    def capture_action(self, action_description):
        """Capture an action and send to callback"""
        current_time = time.time()
        time_since_last = current_time - self.last_action_time
        
        # Add timestamp to action
        timestamp = datetime.now().strftime("%H:%M:%S")
        action_with_time = f"[{timestamp}] {action_description}"
        
        # Log action capture for debugging (if log callback available)
        if self.log_callback and self.root:
            self.root.after(0, lambda: self.log_callback(
                f"Action being captured: {action_description}", "INFO"))
        
        # Send to callback (main thread)
        if self.callback:
            self.root.after(0, lambda: self.callback(action_with_time))
        else:
            # If callback not available, log warning
            if self.log_callback and self.root:
                self.root.after(0, lambda: self.log_callback(
                    f"⚠️ Action captured but callback not available: {action_description}", "WARNING"))
        
        self.last_action_time = current_time
    
    def _get_window_title(self):
        """Get current window title"""
        try:
            if sys.platform == "win32":
                import win32gui
                hwnd = win32gui.GetForegroundWindow()
                return win32gui.GetWindowText(hwnd)
        except:
            pass
        return ""
    
    def _check_tab_switch_after_click(self):
        """Check if a tab switch occurred after a click by monitoring window title changes"""
        try:
            current_time = time.time()
            # Throttle checks to avoid too frequent title reads
            if current_time - self.last_title_check_time < 0.5:
                return
            self.last_title_check_time = current_time
            
            current_title = self._get_window_title()
            if not current_title:
                return
            
            # Only check if we're on target application
            if not self._check_target_application():
                return
            
            # Check if title changed (likely a tab switch if URL is same)
            if self.last_window_title and self.last_window_title != current_title:
                # Check if URL is same (indicates tab switch, not page navigation)
                current_url = None
                if self.browser_monitor and self.browser_monitor.current_url:
                    current_url = self.browser_monitor.current_url
                
                # Extract tab name from title change
                detected_tab = self._extract_tab_name_from_title_change(self.last_window_title, current_title)
                
                if detected_tab:
                    # Update last title
                    self.last_window_title = current_title
                    
                    # Notify callback about tab switch
                    if self.callback and self.root:
                        timestamp = datetime.now().strftime("%H:%M:%S")
                        action = f"Switched to '{detected_tab}' tab"
                        self.root.after(0, lambda: self.callback(f"[{timestamp}] {action}"))
                        
                        # Also log it prominently with tab name clearly displayed
                        if self.log_callback:
                            self.root.after(0, lambda: self.log_callback(
                                "=" * 60, "INFO"))
                            self.root.after(0, lambda: self.log_callback(
                                f"🔄 TAB SWITCH AUTO-DETECTED: Switched to '{detected_tab}' tab", "ACTION"))
                            self.root.after(0, lambda: self.log_callback(
                                f"Tab Name: {detected_tab}", "INFO"))
                            if current_url:
                                self.root.after(0, lambda: self.log_callback(
                                    f"URL unchanged: {current_url}", "INFO"))
                            self.root.after(0, lambda: self.log_callback(
                                f"Previous title: {self.last_window_title}", "INFO"))
                            self.root.after(0, lambda: self.log_callback(
                                f"New title: {current_title}", "INFO"))
                            self.root.after(0, lambda: self.log_callback(
                                f"Tab switch captured automatically", "SUCCESS"))
                            self.root.after(0, lambda: self.log_callback(
                                "=" * 60, "INFO"))
                else:
                    # Update last title even if no tab switch detected
                    self.last_window_title = current_title
            else:
                # Update last title if it's the first time
                if not self.last_window_title:
                    self.last_window_title = current_title
        except Exception as e:
            # Silently handle errors to avoid disrupting click capture
            if self.log_callback and self.root:
                self.root.after(0, lambda: self.log_callback(
                    f"Error checking tab switch: {e}", "ERROR"))
    
    def _extract_tab_name_from_title_change(self, old_title, new_title):
        """Extract tab name from window title change - improved to better detect tab names"""
        try:
            if not old_title or not new_title:
                return None
            
            # Common tab names to look for (case-insensitive)
            common_tab_names = [
                "Accounts", "Users", "Notification", "Settings", "Dashboard", 
                "Profile", "Preferences", "Options", "Configuration",
                "General", "Advanced", "Security", "Privacy", "About"
            ]
            
            # First, check if any common tab name appears in new title but not in old
            old_title_lower = old_title.lower()
            new_title_lower = new_title.lower()
            
            for tab_name in common_tab_names:
                tab_lower = tab_name.lower()
                # Check if tab name is in new title but not in old (or in a different position)
                if tab_lower in new_title_lower:
                    # Check if it's a new occurrence or different position
                    old_positions = [i for i in range(len(old_title_lower)) if old_title_lower.startswith(tab_lower, i)]
                    new_positions = [i for i in range(len(new_title_lower)) if new_title_lower.startswith(tab_lower, i)]
                    
                    # If tab name appears in new title but not in old, or in different position
                    if not old_positions or (new_positions and new_positions != old_positions):
                        # Verify it's a standalone word (not part of another word)
                        import re
                        pattern = r'\b' + re.escape(tab_name) + r'\b'
                        if re.search(pattern, new_title, re.IGNORECASE):
                            return tab_name
            
            # Common separators in window titles
            separators = [' > ', ' - ', ' | ', ' / ', ' :: ', ' >', ' -', ' |']
            
            # Split titles by common separators
            old_parts = None
            new_parts = None
            used_separator = None
            
            for sep in separators:
                if sep in old_title or sep in new_title:
                    old_parts = old_title.split(sep) if sep in old_title else [old_title]
                    new_parts = new_title.split(sep) if sep in new_title else [new_title]
                    used_separator = sep
                    break
            
            # If no separator found, try to find differences word by word
            if not old_parts or not new_parts:
                old_words = old_title.split()
                new_words = new_title.split()
                
                # Find words that are in new title but not in old
                new_unique_words = [w for w in new_words if w not in old_words]
                if new_unique_words:
                    # Take the last unique word (usually the tab name)
                    potential_tab = new_unique_words[-1]
                    # Clean up common suffixes and browser names
                    potential_tab = potential_tab.rstrip(' - > | /').rstrip(' - Google Chrome').rstrip(' - Chrome').rstrip(' - Edge')
                    if len(potential_tab) > 2 and potential_tab.lower() not in ['chrome', 'edge', 'firefox', 'browser']:
                        return potential_tab
                return None
            
            # Compare parts to find what changed
            if len(old_parts) == len(new_parts):
                # Same structure, find the changed part
                for i, (old_part, new_part) in enumerate(zip(old_parts, new_parts)):
                    old_clean = old_part.strip().rstrip(' - Google Chrome').rstrip(' - Chrome').rstrip(' - Edge')
                    new_clean = new_part.strip().rstrip(' - Google Chrome').rstrip(' - Chrome').rstrip(' - Edge')
                    if old_clean != new_clean:
                        # This part changed - likely the tab name
                        changed_part = new_clean
                        # Clean up common prefixes/suffixes
                        changed_part = changed_part.rstrip(' - > | /')
                        if len(changed_part) > 2 and changed_part.lower() not in ['chrome', 'edge', 'firefox', 'browser']:
                            return changed_part
            elif len(new_parts) > len(old_parts):
                # New title has more parts - the extra part is likely the tab
                extra_parts = new_parts[len(old_parts):]
                if extra_parts:
                    tab_name = extra_parts[0].strip()
                    tab_name = tab_name.rstrip(' - > | /').rstrip(' - Google Chrome').rstrip(' - Chrome').rstrip(' - Edge')
                    if len(tab_name) > 2 and tab_name.lower() not in ['chrome', 'edge', 'firefox', 'browser']:
                        return tab_name
            elif len(old_parts) > len(new_parts):
                # Old title had more parts - check what's different
                for i, new_part in enumerate(new_parts):
                    if i < len(old_parts):
                        old_clean = old_parts[i].strip().rstrip(' - Google Chrome').rstrip(' - Chrome').rstrip(' - Edge')
                        new_clean = new_part.strip().rstrip(' - Google Chrome').rstrip(' - Chrome').rstrip(' - Edge')
                        if old_clean != new_clean:
                            changed_part = new_clean
                        changed_part = changed_part.rstrip(' - > | /')
                        if len(changed_part) > 2 and changed_part.lower() not in ['chrome', 'edge', 'firefox', 'browser']:
                            return changed_part
            
            # Fallback: Look for common tab patterns with regex
            import re
            # Look for patterns like "> TabName" or "- TabName" or capitalized words
            patterns = [
                r'[>\-|]\s*([A-Z][a-zA-Z]+)',  # Capitalized word after separator
                r'([A-Z][a-z]+)\s*[-|>]',  # Capitalized word before separator
                r'([A-Z][a-zA-Z]+)\s*$',  # Capitalized word at end (but not browser name)
            ]
            
            for pattern in patterns:
                matches = re.finditer(pattern, new_title)
                for match in matches:
                    potential_tab = match.group(1)
                    # Check if this word wasn't in old title and is not a browser name
                    if potential_tab.lower() not in ['chrome', 'edge', 'firefox', 'browser', 'google']:
                        if potential_tab not in old_title or old_title.find(potential_tab) != new_title.find(potential_tab):
                            return potential_tab
            
            return None
        except Exception as e:
            return None
    
    def _check_tab_switch_in_title(self, window_title, old_title=None):
        """Check if window title indicates a tab switch - generic approach"""
        try:
            if not window_title or not self._check_target_application():
                return
            
            old_title = old_title or self.last_window_title
            if not old_title:
                return
            
            # Check if URL is same (indicates tab switch, not page navigation)
            current_url = None
            if self.browser_monitor and self.browser_monitor.current_url:
                current_url = self.browser_monitor.current_url
            
            # Extract tab name from title change
            detected_tab = self._extract_tab_name_from_title_change(old_title, window_title)
            
            # If tab switch detected, capture it
            if detected_tab:
                if self.callback and self.root:
                    timestamp = datetime.now().strftime("%H:%M:%S")
                    action = f"Switched to '{detected_tab}' tab"
                    self.root.after(0, lambda: self.callback(f"[{timestamp}] {action}"))
                    
                    # Also log it prominently with tab name clearly displayed
                    if self.log_callback:
                        self.root.after(0, lambda: self.log_callback(
                            "=" * 60, "INFO"))
                        self.root.after(0, lambda: self.log_callback(
                            f"🔄 TAB SWITCH AUTO-DETECTED: Switched to '{detected_tab}' tab", "ACTION"))
                        self.root.after(0, lambda: self.log_callback(
                            f"Tab Name: {detected_tab}", "INFO"))
                        if current_url:
                            self.root.after(0, lambda: self.log_callback(
                                f"URL unchanged: {current_url}", "INFO"))
                        self.root.after(0, lambda: self.log_callback(
                            f"Previous title: {old_title}", "INFO"))
                        self.root.after(0, lambda: self.log_callback(
                            f"New title: {window_title}", "INFO"))
                        self.root.after(0, lambda: self.log_callback(
                            f"Tab switch captured automatically from window title change", "SUCCESS"))
                        self.root.after(0, lambda: self.log_callback(
                            "=" * 60, "INFO"))
        except Exception as e:
            pass  # Silently handle errors


class TestCaseCapture:
    def __init__(self, root):
        self.root = root
        self.root.title("Enhanced Auto Test Case Capture Tool")
        self.root.geometry("1100x900")
        self.root.resizable(True, True)
        self.root.minsize(800, 600)  # Set minimum window size for better usability
        
        # Test cases storage organized by module
        self.test_cases_by_module = {}  # {module: [test_cases]}
        
        # Excel file path
        self.excel_file_path = "Doceree_TC.xlsx"
        
        # Test case counter per module
        self.test_case_counters = {}  # {module: counter}
        
        # Current test session data
        self.current_module = ""
        self.current_page = ""
        self.current_url = ""
        self.current_tab = ""  # Track current tab within page
        self.previous_tab = ""  # Track previous tab for switch detection
        self.current_functionality = ""
        self.current_test_steps = []
        self.current_expected_result = ""
        self.current_actual_result = ""
        
        # Track URL change alerts to prevent duplicates
        self.last_url_change_alert = None  # Track last URL change that showed alert
        
        # Detection state flag
        self._detection_in_progress = False
        self._url_detected_flag = False  # Track if URL was detected (for timeout)
        self.url_change_alert_time = 0  # Timestamp of last alert
        self.url_change_alert_showing = False  # Flag to track if alert dialog is currently open
        self._url_cleared_flag = False  # Flag to prevent browser monitor from re-setting URL after clear
        
        # Browser monitoring - start with no base_url, will be set dynamically
        self.browser_monitor = BrowserMonitor(self.on_url_changed, base_url=None)
        
        # Action monitoring
        self.monitor = ActionMonitor(self.on_action_captured, self.browser_monitor)
        self.monitor.root = self.root
        self.monitoring_active = False
        self.auto_save_enabled = False
        self.auto_save_interval = 5  # Auto-save after 5 actions
        self.manual_url_set = False  # Track if URL was set manually
        
        # Logging system
        self.log_messages = []  # Store log messages
        self.max_log_lines = 100  # Maximum log lines to keep
        
        # Load existing test cases
        self.load_existing_test_cases()
        
        # Create GUI
        self.notebook = None
        self.create_widgets()
        
        # Initialize logging after GUI is created
        self.root.after(100, self._initialize_logging)
    
    def _initialize_logging(self):
        """Initialize logging system"""
        self.log_message("=" * 60, "INFO")
        self.log_message("Test Case Capture Tool Started", "SUCCESS")
        self.log_message(f"Base URL: {BASE_URL}", "INFO")
        self.log_message("Ready to capture test cases", "INFO")
        self.log_message("=" * 60, "INFO")
    
    def load_existing_test_cases(self):
        """Load existing test cases from Excel"""
        try:
            wb = load_workbook(self.excel_file_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                module = sheet_name
                if module not in self.test_cases_by_module:
                    self.test_cases_by_module[module] = []
                    self.test_case_counters[module] = 0
                
                # Read test cases from sheet (skip header row)
                for row in ws.iter_rows(min_row=2, values_only=False):
                    if row[0].value:  # If Test Case ID exists
                        test_case = {
                            "test_id": row[0].value,
                            "test_name": row[1].value or "",
                            "description": row[2].value or "",
                            "preconditions": row[3].value or "",
                            "test_steps": row[4].value or "",
                            "expected_result": row[5].value or "",
                            "actual_result": row[6].value or "",
                            "status": row[7].value or "Not Executed",
                            "priority": row[8].value or "Medium",
                            "module": row[9].value or module,
                            "page": row[10].value or "",
                            "tab": row[11].value if len(row) > 11 else "",
                            "url": row[12].value if len(row) > 12 else "",
                            "created_date": row[13].value if len(row) > 13 and isinstance(row[13].value, str) and ':' in str(row[13].value) else datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        }
                        self.test_cases_by_module[module].append(test_case)
                        # Extract counter from test_id
                        try:
                            if test_case["test_id"]:
                                parts = test_case["test_id"].split('_')
                                if len(parts) > 1:
                                    counter = int(parts[-1])
                                    self.test_case_counters[module] = max(self.test_case_counters[module], counter)
                        except:
                            pass
        except FileNotFoundError:
            pass  # File doesn't exist yet
        except Exception as e:
            print(f"Error loading existing test cases: {e}")
    
    def create_widgets(self):
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Setup Tab
        setup_frame = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(setup_frame, text="Setup")
        self.create_setup_tab(setup_frame)
        
        # Capture Tab
        capture_frame = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(capture_frame, text="Auto Capture")
        self.create_capture_tab(capture_frame)
        
        # Status bar
        self.status_label = ttk.Label(self.root, text="Ready - Browser URL monitoring will start automatically", 
                                      relief=tk.SUNKEN, anchor=tk.W, font=("Arial", 9))
        self.status_label.pack(fill=tk.X, side=tk.BOTTOM)
    
    def create_setup_tab(self, parent):
        """Create the setup tab"""
        # Title
        title_label = ttk.Label(parent, text="Enhanced Test Case Capture", 
                               font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Info
        info_text = ("This tool automatically:\n"
                    "• Monitors browser URLs and navigation\n"
                    "• Identifies modules and pages from URLs\n"
                    "• Captures actions for any URL you set\n"
                    "• Organizes test cases by module in Excel")
        info_label = ttk.Label(parent, text=info_text, font=("Arial", 9), 
                              foreground="gray", justify=tk.LEFT)
        info_label.grid(row=1, column=0, columnspan=2, pady=10, sticky=tk.W)
        
        # Base URL (read-only, informational)
        ttk.Label(parent, text="Base URL:", font=("Arial", 10)).grid(
            row=2, column=0, sticky=tk.W, pady=10)
        self.base_url_label = ttk.Label(parent, text="Will be detected from first URL", 
                             font=("Arial", 10, "italic"), 
                             foreground="gray")
        self.base_url_label.grid(row=2, column=1, sticky=tk.W, pady=10, padx=10)
        
        # Functionality (optional, for manual override)
        ttk.Label(parent, text="Functionality (Optional):", font=("Arial", 10)).grid(
            row=3, column=0, sticky=tk.W, pady=10)
        self.functionality_text = tk.Text(parent, height=3, width=50, font=("Arial", 10))
        self.functionality_text.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=10, padx=10)
        ttk.Label(parent, text="(Will be auto-detected from URL if not provided)", 
                 font=("Arial", 8), foreground="gray").grid(
            row=4, column=1, sticky=tk.W, padx=10)
        
        # Start button
        ttk.Button(parent, text="Start Browser Monitoring & Auto-Capture", 
                  command=self.start_browser_monitoring, width=40).grid(
            row=5, column=0, columnspan=2, pady=20)
        
        parent.columnconfigure(1, weight=1)
    
    def create_capture_tab(self, parent):
        """Create the capture tab"""
        # Create a canvas and scrollbar for scrolling
        canvas = tk.Canvas(parent, highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        def configure_scroll_region(event=None):
            # Update scroll region
            canvas.update_idletasks()
            bbox = canvas.bbox("all")
            if bbox:
                canvas.configure(scrollregion=bbox)
        
        def configure_canvas_width(event):
            # Update canvas width to match scrollable_frame
            canvas_width = event.width
            canvas.itemconfig(canvas_frame, width=canvas_width)
        
        scrollable_frame.bind("<Configure>", configure_scroll_region)
        canvas.bind("<Configure>", configure_canvas_width)
        
        canvas_frame = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Bind mousewheel to canvas
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Use scrollable_frame as the parent for all widgets
        # Title
        title_label = ttk.Label(scrollable_frame, text="Automatic Action & URL Capture", 
                               font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Current URL/Module/Page Info
        info_frame = ttk.LabelFrame(scrollable_frame, text="Current Browser State", padding="10")
        info_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        
        url_info_frame = ttk.Frame(info_frame)
        url_info_frame.pack(fill=tk.X)
        
        self.url_label = ttk.Label(url_info_frame, text="URL: Not detected", font=("Arial", 9))
        self.url_label.pack(side=tk.LEFT, anchor=tk.W)
        
        # Add buttons for URL detection (arranged vertically in a column)
        button_frame = ttk.Frame(url_info_frame)
        button_frame.pack(side=tk.RIGHT, padx=5)
        
        self.detect_url_button = ttk.Button(button_frame, text="Detect URL from Browser", 
                  command=self.detect_url_from_browser, width=22)
        self.detect_url_button.pack(pady=2, padx=2, fill=tk.X)
        ttk.Button(button_frame, text="Paste URL Manually", 
                  command=self._paste_url_from_clipboard, width=22).pack(pady=2, padx=2, fill=tk.X)
        ttk.Button(button_frame, text="Clear URL", 
                  command=self.clear_url, width=22).pack(pady=2, padx=2, fill=tk.X)
        
        self.session_info_label = ttk.Label(info_frame, 
            text="Module: Auto-detected | Page: Auto-detected",
            font=("Arial", 9))
        self.session_info_label.pack(anchor=tk.W, pady=(5, 0))
        
        # Status indicator
        self.url_status_label = ttk.Label(info_frame, 
            text="💡 Tip: Open your browser manually, navigate to the page, then click 'Detect URL from Browser' or use Manual Override below",
            font=("Arial", 8), foreground="blue")
        self.url_status_label.pack(anchor=tk.W, pady=(5, 0))
        
        # Manual Override Section
        override_frame = ttk.LabelFrame(scrollable_frame, text="Manual Override (For Manually Opened Browsers)", padding="10")
        override_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        
        ttk.Label(override_frame, text="URL:", font=("Arial", 9)).grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.manual_url_var = tk.StringVar()
        self.manual_url_entry = ttk.Entry(override_frame, textvariable=self.manual_url_var, width=60, font=("Arial", 9))
        self.manual_url_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        # Try to auto-detect URL from browser, otherwise leave empty
        self._try_auto_detect_initial_url()
        # Allow paste with Ctrl+V
        self.manual_url_entry.bind("<Control-v>", lambda e: self.root.after(10, self.set_manual_url))
        ttk.Button(override_frame, text="Set URL", command=self.set_manual_url, width=12).grid(row=0, column=2, padx=5, pady=5)
        ttk.Label(override_frame, text="💡 Copy URL from browser address bar and paste here (Ctrl+V)", 
                 font=("Arial", 7), foreground="gray").grid(row=0, column=3, sticky=tk.W, padx=5)
        
        ttk.Label(override_frame, text="Module:", font=("Arial", 9)).grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.manual_module_var = tk.StringVar()
        module_combo = ttk.Combobox(override_frame, textvariable=self.manual_module_var, 
                                   values=["Login", "Advertiser Dashboard", "Brand Dashboard", 
                                          "Manage Payments", "Manage Users", "Manage Accounts",
                                          "Target", "Plan", "Activate", "Measure", "General"],
                                   width=57, font=("Arial", 9))
        module_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        ttk.Button(override_frame, text="Set Module", command=self.set_manual_module, width=12).grid(row=1, column=2, padx=5, pady=5)
        
        ttk.Label(override_frame, text="Page:", font=("Arial", 9)).grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.manual_page_var = tk.StringVar()
        self.manual_page_entry = ttk.Entry(override_frame, textvariable=self.manual_page_var, width=60, font=("Arial", 9))
        self.manual_page_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        ttk.Button(override_frame, text="Set Page", command=self.set_manual_page, width=12).grid(row=2, column=2, padx=5, pady=5)
        
        override_frame.columnconfigure(1, weight=1)
        
        ttk.Label(override_frame, text="💡 Workflow: 1) Copy URL from browser → 2) Paste here → 3) Click 'Set URL' → 4) Module/Page auto-detected", 
                 font=("Arial", 7), foreground="gray", justify=tk.LEFT).grid(row=3, column=0, columnspan=4, sticky=tk.W, padx=5, pady=5)
        
        # Monitoring Controls
        control_frame = ttk.LabelFrame(scrollable_frame, text="Monitoring Controls", padding="10")
        control_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        
        self.monitor_status_label = ttk.Label(control_frame, 
            text="Status: Monitoring OFF", 
            font=("Arial", 10, "bold"), foreground="red")
        self.monitor_status_label.pack(side=tk.LEFT, padx=10)
        
        self.start_monitor_btn = ttk.Button(control_frame, text="Start Auto-Capture", 
                                           command=self.start_monitoring, width=20)
        self.start_monitor_btn.pack(side=tk.LEFT, padx=5)
        
        self.stop_monitor_btn = ttk.Button(control_frame, text="Stop Auto-Capture", 
                                          command=self.stop_monitoring, width=20, state=tk.DISABLED)
        self.stop_monitor_btn.pack(side=tk.LEFT, padx=5)
        
        # Auto-save option
        self.auto_save_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(control_frame, text="Auto-save after 5 actions", 
                       variable=self.auto_save_var).pack(side=tk.LEFT, padx=10)
        
        # Captured Actions Section
        actions_frame = ttk.LabelFrame(scrollable_frame, text="Automatically Captured Actions", padding="10")
        actions_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        
        # Actions listbox with scrollbar
        actions_list_frame = ttk.Frame(actions_frame)
        actions_list_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(actions_list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.actions_listbox = tk.Listbox(actions_list_frame, height=8, font=("Arial", 9),
                                         yscrollcommand=scrollbar.set)
        self.actions_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.actions_listbox.yview)
        
        # Action count
        self.action_count_label = ttk.Label(actions_frame, text="Actions captured: 0", 
                                            font=("Arial", 9))
        self.action_count_label.pack(pady=5)
        
        # Manual action entry (optional)
        manual_frame = ttk.Frame(actions_frame)
        manual_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(manual_frame, text="Add Manual Action:", font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        self.manual_action_entry = ttk.Entry(manual_frame, width=40, font=("Arial", 9))
        self.manual_action_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        self.manual_action_entry.bind("<Return>", lambda e: self.add_manual_action())
        
        ttk.Button(manual_frame, text="Add", 
                  command=self.add_manual_action, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(manual_frame, text="Remove Selected", 
                  command=self.remove_action, width=15).pack(side=tk.LEFT, padx=5)
        
        # Quick action templates
        template_frame = ttk.Frame(actions_frame)
        template_frame.pack(fill=tk.X, pady=5)
        ttk.Label(template_frame, text="Quick Templates:", font=("Arial", 8), foreground="gray").pack(side=tk.LEFT, padx=5)
        ttk.Button(template_frame, text="Navigate to page", 
                  command=lambda: self.add_template_action("Navigate to page"), width=16).pack(side=tk.LEFT, padx=2)
        ttk.Button(template_frame, text="Click button", 
                  command=lambda: self.add_template_action("Click on button"), width=14).pack(side=tk.LEFT, padx=2)
        ttk.Button(template_frame, text="Open dropdown", 
                  command=lambda: self.add_template_action("Open dropdown menu"), width=14).pack(side=tk.LEFT, padx=2)
        ttk.Button(template_frame, text="Select from menu", 
                  command=lambda: self.add_template_action("Select option from menu"), width=16).pack(side=tk.LEFT, padx=2)
        ttk.Button(template_frame, text="Enter text", 
                  command=lambda: self.add_template_action("Enter text in field"), width=14).pack(side=tk.LEFT, padx=2)
        ttk.Button(template_frame, text="Switch Tab", 
                  command=self.capture_tab_switch, width=14).pack(side=tk.LEFT, padx=2)
        ttk.Button(template_frame, text="Verify element", 
                  command=lambda: self.add_template_action("Verify element is displayed"), width=16).pack(side=tk.LEFT, padx=2)
        
        # Current Tab Tracker
        tab_tracker_frame = ttk.Frame(actions_frame)
        tab_tracker_frame.pack(fill=tk.X, pady=5)
        ttk.Label(tab_tracker_frame, text="Current Tab:", font=("Arial", 8), foreground="gray").pack(side=tk.LEFT, padx=5)
        self.current_tab_var = tk.StringVar()
        self.current_tab_entry = ttk.Entry(tab_tracker_frame, textvariable=self.current_tab_var, width=30, font=("Arial", 8))
        self.current_tab_entry.pack(side=tk.LEFT, padx=5)
        self.current_tab_entry.bind("<Return>", lambda e: self.update_current_tab())
        ttk.Button(tab_tracker_frame, text="Update Tab", 
                  command=self.update_current_tab, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Label(tab_tracker_frame, text="💡 Enter tab name (e.g., Accounts, Users, Notification) and click 'Switch Tab' when switching", 
                 font=("Arial", 7), foreground="gray").pack(side=tk.LEFT, padx=5)
        
        # Expected Result
        ttk.Label(scrollable_frame, text="Expected Result:", font=("Arial", 10)).grid(
            row=5, column=0, sticky=tk.W, pady=5)
        self.expected_result_text = tk.Text(scrollable_frame, height=2, width=50, font=("Arial", 10))
        self.expected_result_text.grid(row=5, column=1, sticky=(tk.W, tk.E), pady=5, padx=10)
        
        # Actual Result
        ttk.Label(scrollable_frame, text="Actual Result:", font=("Arial", 10)).grid(
            row=6, column=0, sticky=tk.W, pady=5)
        self.actual_result_text = tk.Text(scrollable_frame, height=2, width=50, font=("Arial", 10))
        self.actual_result_text.grid(row=6, column=1, sticky=(tk.W, tk.E), pady=5, padx=10)
        
        # Status
        ttk.Label(scrollable_frame, text="Status:", font=("Arial", 10)).grid(
            row=7, column=0, sticky=tk.W, pady=5)
        self.status_var = tk.StringVar(value="Not Executed")
        status_combo = ttk.Combobox(scrollable_frame, textvariable=self.status_var, 
                                   values=["Not Executed", "Pass", "Fail", "Blocked"],
                                   state="readonly", width=47)
        status_combo.grid(row=7, column=1, sticky=tk.W, pady=5, padx=10)
        
        # Buttons
        button_frame = ttk.Frame(scrollable_frame)
        button_frame.grid(row=8, column=0, columnspan=2, pady=20)
        
        ttk.Button(button_frame, text="Save Test Case to Excel", 
                  command=self.save_test_case, width=25).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear Actions", 
                  command=self.clear_actions, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear Logs", 
                  command=self.clear_logs, width=15).pack(side=tk.LEFT, padx=5)
        
        # Activity Log Section - Make it larger and resizable
        log_frame = ttk.LabelFrame(scrollable_frame, text="📊 Activity Log & Processing Status (Resizable - Drag window to resize)", padding="10")
        log_frame.grid(row=9, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        log_frame.grid_rowconfigure(0, weight=1, minsize=200)  # Minimum height of 200px, but can expand
        log_frame.grid_columnconfigure(0, weight=1)
        
        # Log text widget with scrollbar - Make it expandable
        log_text_frame = ttk.Frame(log_frame)
        log_text_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        log_text_frame.grid_rowconfigure(0, weight=1)
        log_text_frame.grid_columnconfigure(0, weight=1)
        
        log_scrollbar = ttk.Scrollbar(log_text_frame)
        log_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Remove fixed height - let it expand based on available space
        self.log_text = tk.Text(log_text_frame, font=("Consolas", 8), 
                               yscrollcommand=log_scrollbar.set, wrap=tk.WORD,
                               bg="#1e1e1e", fg="#d4d4d4", insertbackground="white")
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        log_scrollbar.config(command=self.log_text.yview)
        
        # Add resize handle hint
        resize_hint = ttk.Label(log_frame, text="💡 Tip: Resize the main window to make this log area taller or shorter", 
                               font=("Arial", 7), foreground="gray")
        resize_hint.grid(row=1, column=0, sticky=tk.W, pady=(5, 0))
        
        # Configure text tags for different log levels
        self.log_text.tag_config("INFO", foreground="#4ec9b0")
        self.log_text.tag_config("SUCCESS", foreground="#4ec9b0", font=("Consolas", 8, "bold"))
        self.log_text.tag_config("WARNING", foreground="#dcdcaa")
        self.log_text.tag_config("ERROR", foreground="#f48771")
        self.log_text.tag_config("ACTION", foreground="#569cd6")
        self.log_text.tag_config("URL", foreground="#ce9178")
        self.log_text.tag_config("TIMESTAMP", foreground="#808080")
        
        # Configure grid weights for proper expansion
        scrollable_frame.columnconfigure(0, weight=1)
        scrollable_frame.columnconfigure(1, weight=1)
        scrollable_frame.rowconfigure(9, weight=3)  # Give more weight to log section - it will expand when window is resized
    
    def start_browser_monitoring(self):
        """Start browser URL monitoring"""
        self.log_message("Starting browser monitoring...", "INFO")
        if self.browser_monitor.start_monitoring():
            self.log_message("Browser monitoring started successfully", "SUCCESS")
            self.status_label.config(text="Browser monitoring started! You can open browser manually and use 'Detect URL from Browser' button.")
            self.log_message("Ready to detect URLs. Open browser and navigate to target application", "INFO")
            
            # Try to auto-detect URL after monitoring starts (non-blocking)
            self.root.after(1000, self._try_auto_detect_url_after_monitoring_start)
            self.notebook.select(1)  # Switch to capture tab
        else:
            self.log_message("Failed to start browser monitoring", "ERROR")
            messagebox.showerror("Error", "Failed to start browser monitoring!")
    
    def on_url_changed(self, action, url, module, page):
        """Callback when URL changes - automatically updates URL without popup"""
        # If URL was just cleared, ignore auto-updates for a short time
        if hasattr(self, '_url_cleared_flag') and self._url_cleared_flag:
            # URL was cleared - don't auto-update for 2 seconds
            import time
            if hasattr(self, '_url_clear_time'):
                elapsed = time.time() - self._url_clear_time
                if elapsed < 2.0:  # Ignore updates for 2 seconds after clear
                    self.log_message(f"⏸️ Ignoring auto URL update (URL was just cleared): {url}", "INFO")
                    return
        
        old_url = self.current_url
        
        # Prevent duplicate processing of the same URL change
        if url == self.current_url:
            return  # URL hasn't changed, skip processing
        
        # Log URL detection prominently
        self.log_message("=" * 60, "INFO")
        self.log_message(f"🌐 URL CHANGE DETECTED: {url}", "URL")
        if old_url:
            self.log_message(f"Previous URL: {old_url}", "INFO")
        self.log_message(f"New URL: {url}", "SUCCESS")
        self.log_message(f"Module identified: {module} | Page: {page}", "SUCCESS")
        self.log_message("URL updated automatically - monitoring continues", "INFO")
        self.log_message("=" * 60, "INFO")
        
        # Always update URL automatically without showing popup
        self.current_url = url
        self.current_module = module
        self.current_page = page
        
        # Update UI
        self.root.after(0, lambda: self.update_url_info(url, module, page))
        
        # Add navigation action to list
        self.root.after(0, lambda: self.add_navigation_action(action))
    
    def _show_url_change_alert(self, old_url, new_url, module, page):
        """Show alert when URL changes and ask user what to do"""
        try:
            response = messagebox.askyesno("URL Changed Detected", 
                                          f"The URL in your browser has changed!\n\n"
                                          f"Previous URL: {old_url}\n"
                                          f"New URL: {new_url}\n\n"
                                          f"What do you want to do?\n\n"
                                          f"• Click 'Yes' to update to the new URL and continue\n"
                                          f"• Click 'No' to keep the existing URL")
            if response:
                # User wants to update to new URL
                self.current_url = new_url
                self.current_module = module
                self.current_page = page
                
                # Update UI
                self.update_url_info(new_url, module, page)
                
                # Add navigation action to list
                self.add_navigation_action(f"Navigated to: {new_url}\nModule: {module} | Page: {page}")
                
                self.log_message(f"✅ URL updated to: {new_url}", "SUCCESS")
            else:
                # User wants to keep existing URL - don't update
                self.log_message(f"User chose to keep existing URL: {old_url}", "INFO")
                # Don't update the URL, but the browser_monitor already has the new URL
                # We need to revert it back
                self.browser_monitor.current_url = old_url
                # Revert module and page too if needed
                if self.current_module and self.current_page:
                    self.browser_monitor.current_module = self.current_module
                    self.browser_monitor.current_page = self.current_page
        finally:
            # Always reset the flag when alert is closed
            self.url_change_alert_showing = False
    
    def _try_auto_detect_initial_url(self):
        """Try to auto-detect URL from browser when tool starts"""
        try:
            # Try to get URL from Chrome DevTools
            url = self.browser_monitor._get_url_from_chrome_devtools()
            if url:
                self.manual_url_var.set(url)  # Set full URL
                self.log_message(f"Auto-detected URL from browser: {url}", "INFO")
            return
        
            # Try window title as fallback
            if sys.platform == "win32":
                try:
                    import win32gui
                    hwnd = win32gui.GetForegroundWindow()
                    window_title = win32gui.GetWindowText(hwnd)
                    if window_title:
                        # Try to extract URL from title
                        import re
                        url_pattern = r'https?://[^\s]+'
                        match = re.search(url_pattern, window_title)
                        if match:
                            url = match.group(0)
                            self.manual_url_var.set(url)
                            self.log_message(f"Auto-detected URL from window title: {url}", "INFO")
                except:
                    pass
        except:
            pass
        
        # If no URL detected, leave field empty (dynamic - no hardcoded value)
        self.manual_url_var.set("")
    
    def _try_auto_detect_url_after_monitoring_start(self):
        """Try to auto-detect URL after browser monitoring starts"""
        try:
            url = self.browser_monitor._get_url_from_chrome_devtools()
            if url:
                self.manual_url_var.set(url)
                # Also update the current URL if it matches
                if not self.current_url or url != self.current_url:
                    self.browser_monitor._handle_url_change(url)
                self.log_message(f"Auto-detected URL after monitoring start: {url}", "SUCCESS")
        except:
            pass
    
    def update_url_info(self, url, module, page):
        """Update URL info in UI"""
        self.url_label.config(text=f"URL: {url}")
        self.session_info_label.config(text=f"Module: {module} | Page: {page}")
        self.status_label.config(text=f"Monitoring: {module} - {page}")
        base_url = self.browser_monitor.base_url if self.browser_monitor else BASE_URL
        if url and (not base_url or url.startswith(base_url)):
            self.url_status_label.config(text="✓ URL detected and monitoring active", foreground="green")
            # Update base URL label if it was just set
            if hasattr(self, 'base_url_label') and base_url:
                self.base_url_label.config(text=base_url, foreground="blue", font=("Arial", 10, "bold"))
        else:
            self.url_status_label.config(text="⚠ URL not detected. Use 'Detect URL from Browser' or Manual Override", foreground="orange")
    
    def detect_url_from_browser(self):
        """Detect URL from any open browser window - runs in background thread for fast response"""
        # FORCE RESET: Check if button is stuck in "Detecting..." state and reset if needed
        if hasattr(self, 'detect_url_button'):
            try:
                button_text = self.detect_url_button.cget('text')
                button_state = self.detect_url_button.cget('state')
                # If button is disabled with "Detecting..." text, force reset
                if button_state == 'disabled' and 'Detecting' in button_text:
                    self.log_message("🔧 Button was stuck - forcing reset...", "WARNING")
                    self._detection_in_progress = False
                    self._reset_detection_button()
                    self._hide_loading_indicator()
                    # Small delay to ensure reset completes
                    import time
                    time.sleep(0.1)
            except:
                pass
        
        # Prevent multiple simultaneous detections (but allow if we just reset)
        if hasattr(self, '_detection_in_progress') and self._detection_in_progress:
            self.log_message("⚠️ Detection already in progress. Please wait...", "WARNING")
            return
        
        # First, check if any browser is open
        browser_found = self._check_if_browser_is_open()
        if not browser_found:
            messagebox.showwarning("No Browser Found", 
                                 "No browser window detected!\n\n"
                                 "Please:\n"
                                 "1. Open your browser (Chrome, Edge, or Firefox)\n"
                                 "2. Navigate to the page you want to test\n"
                                 "3. Make sure the browser window is visible (not minimized)\n"
                                 "4. Try 'Detect URL from Browser' again\n\n"
                                 "Or use 'Paste URL Manually' button to enter URL directly.")
            return
        
        # Set detection flag
        self._detection_in_progress = True
        
        # Track if URL was detected (for timeout check)
        self._url_detected_flag = False
        
        # Reset URL cleared flag (allow detection to work)
        self._url_cleared_flag = False
        if hasattr(self, '_url_clear_time'):
            delattr(self, '_url_clear_time')
        
        # Store detection start time for timeout tracking
        import time
        self._detection_start_time = time.time()
        
        # Disable button and show loader
        if hasattr(self, 'detect_url_button'):
            try:
                self.detect_url_button.config(state='disabled', text='⏳ Detecting...')
            except:
                # If button doesn't exist or is destroyed, continue anyway
                pass
        
        # Show loading indicator
        self._show_loading_indicator()
        self.log_message("Detecting URL from browser...", "INFO")
        self.log_message(f"Browser windows found: {browser_found}", "INFO")
        
        # Run detection in background thread to avoid blocking UI
        def detect_in_thread():
            try:
                url = None
                detection_details = []
                
                # First, try to find and activate a browser window (with timeout)
                # (The tool window might be active when button is clicked)
                try:
                    browser_hwnd = self._find_and_activate_any_browser_window()
                    if browser_hwnd:
                        self.root.after(0, lambda: self.log_message("✅ Browser window found and activated", "SUCCESS"))
                        import time
                        time.sleep(0.2)  # Reduced wait time
                    else:
                        # Check if current active window is a browser
                        active_browser = self._verify_active_window_is_browser()
                        if not active_browser:
                            detection_details.append("⚠️ No browser window found - please open a browser first")
                            self.root.after(0, lambda: self.log_message("⚠️ No browser window found. Please open your browser first.", "WARNING"))
                        else:
                            self.root.after(0, lambda: self.log_message(f"✅ Active window is a browser: {active_browser}", "INFO"))
                except Exception as e:
                    detection_details.append(f"Window activation error: {str(e)}")
                    self.root.after(0, lambda: self.log_message(f"⚠️ Error activating browser window: {str(e)}", "WARNING"))
                
                # Log detection state for debugging
                self.root.after(0, lambda: self.log_message(f"🔍 Detection state: _detection_in_progress={getattr(self, '_detection_in_progress', 'NOT SET')}", "INFO"))
                
                # Method 1: Try keyboard automation first (most reliable - works for all browsers)
                # This should work even if window title doesn't have URL
                self.root.after(0, lambda: self.log_message("Method 1: Trying keyboard automation (Ctrl+L, Ctrl+C)...", "INFO"))
                try:
                    url = self._try_get_url_via_keyboard()
                    if url:
                        self.root.after(0, lambda: self.log_message(f"✅ URL found via keyboard: {url}", "SUCCESS"))
                    else:
                        detection_details.append("Keyboard: Could not copy URL from address bar (make sure browser window is active)")
                except Exception as e:
                    detection_details.append(f"Keyboard error: {str(e)}")
                
                if not url:
                    # Method 2: Try Chrome DevTools Protocol (works for Chrome and Edge - ~0.2s per port)
                    self.root.after(0, lambda: self.log_message("Method 2: Trying Chrome DevTools Protocol...", "INFO"))
                    try:
                        url = self._get_url_from_chrome_devtools_simple()
                        if url:
                            self.root.after(0, lambda: self.log_message(f"✅ URL found via DevTools: {url}", "SUCCESS"))
                        else:
                            detection_details.append("DevTools: No browser with remote debugging found (ports 9222-9226)")
                    except Exception as e:
                        detection_details.append(f"DevTools error: {str(e)}")
                
                if not url:
                    # Method 3: Try window title detection (many browsers don't show URL in title)
                    self.root.after(0, lambda: self.log_message("Method 3: Trying window title detection...", "INFO"))
                    try:
                        url = self._get_url_from_window_title_simple()
                        if url:
                            self.root.after(0, lambda: self.log_message(f"✅ URL found via window title: {url}", "SUCCESS"))
                        else:
                            detection_details.append("Window title: No URL found in browser window title (this is normal for most browsers)")
                    except Exception as e:
                        detection_details.append(f"Window title error: {str(e)}")
                
                # Log detection details if all methods failed
                if not url and detection_details:
                    self.root.after(0, lambda: self.log_message("=" * 60, "INFO"))
                    self.root.after(0, lambda: self.log_message("❌ All detection methods failed:", "WARNING"))
                    for detail in detection_details:
                        self.root.after(0, lambda d=detail: self.log_message(f"  • {d}", "WARNING"))
                    self.root.after(0, lambda: self.log_message("=" * 60, "INFO"))
                
                # Mark URL as detected if found
                if url:
                    self._url_detected_flag = True
                    # Reset detection flag since we got a result
                    self._detection_in_progress = False
                
                # Update UI in main thread
                # If URL is None, don't reset detection_in_progress - let timeout handle it
                if url:
                    # URL found - reset flag immediately
                    self._detection_in_progress = False
                # else: keep flag True so timeout can detect it
                
                self.root.after(0, lambda: self._handle_detection_result(url))
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                # Don't reset flag here - let timeout handle it
                self.root.after(0, lambda: self._handle_detection_error(f"{str(e)}\n\nDetails:\n{error_details}"))
        
        # Start detection thread
        thread = threading.Thread(target=detect_in_thread, daemon=True)
        thread.start()
        
        # 10-second timeout: Check if URL was detected, show alert if not
        def check_detection_timeout():
            try:
                import time
                # Check if enough time has passed (at least 10 seconds since detection started)
                if hasattr(self, '_detection_start_time'):
                    elapsed = time.time() - self._detection_start_time
                    if elapsed >= 10:
                        # Check if URL was detected
                        url_was_detected = hasattr(self, '_url_detected_flag') and self._url_detected_flag
                        
                        # If URL wasn't detected, show alert (regardless of detection_in_progress state)
                        if not url_was_detected:
                            # Check if button is still in detecting state
                            button_still_detecting = False
                            if hasattr(self, 'detect_url_button'):
                                try:
                                    button_state = self.detect_url_button.cget('state')
                                    button_text = self.detect_url_button.cget('text')
                                    if button_state == 'disabled' and 'Detecting' in button_text:
                                        button_still_detecting = True
                                except:
                                    pass
                            
                            # URL not detected within 10 seconds - show alert
                            self.log_message("⚠️ URL detection timeout (10s) - URL not detected", "WARNING")
                            self._detection_in_progress = False
                            self._reset_detection_button()
                            self._hide_loading_indicator()
                            
                            # Only show alert if button is still detecting (to avoid duplicate alerts)
                            if button_still_detecting:
                                # Show alert dialog
                                messagebox.showwarning("URL Not Detected", 
                                                     "URL could not be detected within 10 seconds.\n\n"
                                                     "Possible reasons:\n"
                                                     "• Browser window is not active/focused\n"
                                                     "• Browser doesn't have remote debugging enabled\n"
                                                     "• URL is not accessible from current methods\n\n"
                                                     "Please try:\n"
                                                     "1. Click on your browser window to make it active\n"
                                                     "2. Use 'Paste URL Manually' button (Ctrl+L, Ctrl+A, Ctrl+C in browser)\n"
                                                     "3. Or use 'Manual Override' section to enter URL directly")
                                self.log_message("💡 Tip: Use 'Paste URL Manually' for faster detection", "INFO")
                        else:
                            # URL was detected, just make sure button is reset
                            if hasattr(self, '_detection_in_progress') and self._detection_in_progress:
                                self._detection_in_progress = False
                                self._reset_detection_button()
                                self._hide_loading_indicator()
            except Exception as e:
                # Even if there's an error, try to reset
                try:
                    self._detection_in_progress = False
                    self._reset_detection_button()
                    self._hide_loading_indicator()
                except:
                    pass
        
        # Safety timeout: Reset button state after 3 seconds if still detecting (very aggressive)
        def safety_timeout():
            try:
                if hasattr(self, '_detection_in_progress') and self._detection_in_progress:
                    # Don't reset here if we're waiting for the 10-second timeout
                    # Just log a warning
                    pass
                # Also check button state directly
                if hasattr(self, 'detect_url_button'):
                    try:
                        button_state = self.detect_url_button.cget('state')
                        button_text = self.detect_url_button.cget('text')
                        if button_state == 'disabled' and 'Detecting' in button_text:
                            # Only reset if URL was already detected or if it's been too long
                            if hasattr(self, '_url_detected_flag') and self._url_detected_flag:
                                self.log_message("🔧 Button still stuck after detection - forcing reset", "WARNING")
                                self._detection_in_progress = False
                                self._reset_detection_button()
                                self._hide_loading_indicator()
                    except:
                        pass
            except:
                pass
        
        self.root.after(3000, safety_timeout)  # 3 second safety check
        self.root.after(10000, check_detection_timeout)  # 10 second timeout with alert
    
    def _show_loading_indicator(self):
        """Show animated loading indicator"""
        # Stop any existing animation
        if hasattr(self, 'loading_animation_id') and self.loading_animation_id is not None:
            try:
                self.root.after_cancel(self.loading_animation_id)
            except (tk.TclError, ValueError):
                # Invalid ID - ignore
                pass
            self.loading_animation_id = None
        
        if not hasattr(self, 'loading_label'):
            # Create loading label if it doesn't exist
            if hasattr(self, 'url_status_label'):
                # Place it near the status label
                self.loading_label = ttk.Label(self.url_status_label.master, 
                                              text="⏳ Detecting URL...", 
                                              font=("Arial", 9, "bold"),
                                              foreground="blue")
                self.loading_label.pack(anchor=tk.W, pady=(2, 0))
            else:
                return
        
        # Show the label
        self.loading_label.pack(anchor=tk.W, pady=(2, 0))
        
        # Update loading text with animation
        self.loading_dots = 0
        self.loading_active = True
        self._animate_loading()
    
    def _animate_loading(self):
        """Animate loading indicator"""
        if not hasattr(self, 'loading_active') or not self.loading_active:
            return
        
        if hasattr(self, 'loading_label') and self.loading_label.winfo_exists():
            dots = "." * (self.loading_dots % 4)
            self.loading_label.config(text=f"⏳ Detecting URL{dots}")
            self.loading_dots += 1
            # Continue animation until detection completes
            self.loading_animation_id = self.root.after(300, self._animate_loading)
    
    def _hide_loading_indicator(self):
        """Hide loading indicator"""
        # Stop animation
        self.loading_active = False
        if hasattr(self, 'loading_animation_id') and self.loading_animation_id is not None:
            try:
                self.root.after_cancel(self.loading_animation_id)
            except (tk.TclError, ValueError):
                # Invalid ID - ignore
                pass
            self.loading_animation_id = None
        
        # Hide the label
        if hasattr(self, 'loading_label') and self.loading_label.winfo_exists():
            self.loading_label.pack_forget()
    
    def _check_if_browser_is_open(self):
        """Check if any browser window is open and return count"""
        try:
            if sys.platform == "win32":
                import win32gui
                browser_keywords = ['chrome', 'edge', 'firefox', 'opera', 'brave', 'vivaldi', 'microsoft edge']
                browser_count = 0
                browser_windows = []
                
                def enum_windows_callback(hwnd, param):
                    try:
                        if win32gui.IsWindowVisible(hwnd):
                            window_title = win32gui.GetWindowText(hwnd)
                            if window_title:
                                title_lower = window_title.lower()
                                if any(keyword in title_lower for keyword in browser_keywords):
                                    # Exclude extension windows and popups
                                    if 'extension' not in title_lower and 'popup' not in title_lower:
                                        param.append(window_title)
                    except:
                        pass
                    return True
                
                win32gui.EnumWindows(enum_windows_callback, browser_windows)
                browser_count = len(browser_windows)
                
                if browser_count > 0:
                    self.log_message(f"Found {browser_count} browser window(s): {', '.join(browser_windows[:3])}", "INFO")
                else:
                    self.log_message("No browser windows found", "WARNING")
                
                return browser_count
        except:
            pass
        return 0
    
    def _verify_active_window_is_browser(self):
        """Verify that the currently active window is a browser - returns browser name or None"""
        try:
            if sys.platform == "win32":
                import win32gui
                browser_keywords = ['chrome', 'edge', 'firefox', 'opera', 'brave', 'vivaldi', 'microsoft edge']
                
                active_hwnd = win32gui.GetForegroundWindow()
                if not active_hwnd:
                    return None
                
                active_window_title = win32gui.GetWindowText(active_hwnd)
                if not active_window_title:
                    return None
                
                title_lower = active_window_title.lower()
                
                # Check if active window is a browser
                for keyword in browser_keywords:
                    if keyword in title_lower:
                        # Exclude extension windows and popups
                        if 'extension' not in title_lower and 'popup' not in title_lower:
                            return keyword.title()  # Return browser name
                
                return None
        except:
            pass
        return None
    
    def _reset_detection_button(self):
        """Reset detection button state - helper method"""
        try:
            if hasattr(self, 'detect_url_button'):
                # Check if widget exists before trying to configure it
                try:
                    # Try to get widget info - if this fails, widget doesn't exist
                    _ = self.detect_url_button.winfo_exists()
                    # Force reset both state and text
                    self.detect_url_button.config(state='normal', text='Detect URL from Browser')
                    # Verify it worked
                    current_state = self.detect_url_button.cget('state')
                    if current_state != 'normal':
                        # Force it again
                        self.detect_url_button.config(state='normal')
                except tk.TclError:
                    # Widget was destroyed, that's OK
                    pass
        except Exception as e:
            # Silently fail - button might not exist yet
            pass
    
    def _handle_detection_result(self, url):
        """Handle detection result in main thread"""
        # Mark URL as detected only if URL is not None
        if url:
            self._url_detected_flag = True
            # Reset detection flag IMMEDIATELY when URL is found
            self._detection_in_progress = False
        # else: Keep flag True so timeout can detect it
        
        # Hide loading indicator
        self._hide_loading_indicator()
        
        # Re-enable button IMMEDIATELY
        self._reset_detection_button()
        
        # Double-check button is reset (safety)
        if hasattr(self, 'detect_url_button'):
            try:
                button_state = self.detect_url_button.cget('state')
                if button_state == 'disabled':
                    self.log_message("🔧 Button still disabled - forcing reset", "WARNING")
                    self._reset_detection_button()
            except:
                pass
        
        if url:
            self.log_message(f"✅ URL detected: {url}", "SUCCESS")
            # Check if it's different from current URL
            if url != self.current_url:
                # URL changed - ask user what to do
                self._handle_url_change_with_confirmation(url, None, None)
            else:
                # Same URL - just confirm
                self.log_message(f"✅ Current URL confirmed: {url}", "SUCCESS")
                messagebox.showinfo("URL Detected", 
                                  f"URL detected from browser:\n{url}\n\n"
                                  f"Module: {self.current_module}\n"
                                  f"Page: {self.current_page}\n\n"
                                  f"This URL is already set in the tool.")
        else:
            # Could not detect URL
            self.log_message("❌ Could not detect URL from browser", "WARNING")
            self._show_detection_failed_dialog()
    
    def _handle_detection_error(self, error_msg):
        """Handle detection error in main thread"""
        # Don't reset flag here - let the 10-second timeout handle it
        # This ensures the timeout alert will show if detection fails
        # The flag will be reset by the timeout function
        
        # Hide loading indicator
        self._hide_loading_indicator()
        
        # Re-enable button IMMEDIATELY
        self._reset_detection_button()
        
        # Double-check button is reset (safety)
        if hasattr(self, 'detect_url_button'):
            try:
                button_state = self.detect_url_button.cget('state')
                if button_state == 'disabled':
                    self.log_message("🔧 Button still disabled after error - forcing reset", "WARNING")
                    self._reset_detection_button()
            except:
                pass
        
        self.log_message(f"❌ Error during URL detection: {error_msg}", "ERROR")
        self._show_detection_failed_dialog()
    
    def _get_url_from_chrome_devtools_simple(self):
        """Simple Chrome DevTools URL detection - checks active browser window or finds one"""
        try:
            import json as json_lib
            import urllib.request
            import urllib.error
            import win32gui
            
            # Get the active/foreground window
            active_hwnd = win32gui.GetForegroundWindow()
            active_window_title = None
            
            if active_hwnd:
                active_window_title = win32gui.GetWindowText(active_hwnd)
            
            # Verify it's a browser window and identify browser type
            browser_keywords = ['chrome', 'edge', 'firefox', 'opera', 'brave', 'vivaldi', 'microsoft edge']
            active_browser_type = None
            
            if active_window_title:
                title_lower = active_window_title.lower()
                for keyword in browser_keywords:
                    if keyword in title_lower:
                        # Exclude extension windows and popups
                        if 'extension' not in title_lower and 'popup' not in title_lower:
                            active_browser_type = keyword
                            break
            
            # If active window is not a browser, try to find one
            if not active_browser_type:
                browser_hwnd = self._find_and_activate_any_browser_window()
                if browser_hwnd:
                    active_hwnd = browser_hwnd
                    active_window_title = win32gui.GetWindowText(active_hwnd)
                    if active_window_title:
                        title_lower = active_window_title.lower()
                        for keyword in browser_keywords:
                            if keyword in title_lower:
                                if 'extension' not in title_lower and 'popup' not in title_lower:
                                    active_browser_type = keyword
                                    break
                else:
                    return None  # No browser window found
            
            # Firefox doesn't support Chrome DevTools Protocol - skip it
            if active_browser_type and 'firefox' in active_browser_type:
                return None
            
            # Try multiple common ports with reduced timeout for faster response
            ports = [9222, 9223, 9224, 9225, 9226]
            for port in ports:
                try:
                    # Reduced timeout to 0.2s for even faster response
                    response = urllib.request.urlopen(f"http://localhost:{port}/json", timeout=0.2)
                    tabs = json_lib.loads(response.read().decode())
                    
                    if not tabs:
                        continue
                    
                    # STRICT: Only match tabs that match the active window title
                    # This ensures we only get URLs from the active browser window
                    best_match = None
                    best_match_score = 0
                    
                    for tab in tabs:
                        tab_title = tab.get('title', '')
                        url = tab.get('url', '')
                        
                        if not url or not (url.startswith('http://') or url.startswith('https://')):
                            continue
                        
                        # Skip chrome:// and about: pages
                        if url.startswith('chrome://') or url.startswith('about:'):
                            continue
                        
                        if tab_title and active_window_title:
                            tab_title_lower = tab_title.lower()
                            active_title_lower = active_window_title.lower()
                            
                            # Calculate match score - higher is better
                            score = 0
                            
                            # Exact match gets highest score
                            if tab_title_lower == active_title_lower:
                                score = 100
                            # Tab title contained in window title
                            elif tab_title_lower in active_title_lower:
                                score = 80
                            # Window title contained in tab title
                            elif active_title_lower in tab_title_lower:
                                score = 70
                            # Word matching
                            else:
                                tab_words = set(word.lower() for word in tab_title_lower.split() if len(word) > 3)
                                active_words = set(word.lower() for word in active_title_lower.split() if len(word) > 3)
                                common_words = tab_words.intersection(active_words)
                                if common_words:
                                    score = len(common_words) * 10
                            
                            if score > best_match_score:
                                best_match = url
                                best_match_score = score
                    
                    # Return best match if found (lowered threshold for better detection)
                    if best_match and best_match_score >= 30:  # Lowered threshold from 50 to 30
                        return best_match
                    
                    # If no good match but we have tabs, return the first valid URL from active tab
                    # This is a fallback - the active tab usually has webSocketDebuggerUrl
                    if not best_match:
                        for tab in tabs:
                            if 'webSocketDebuggerUrl' in tab and tab['webSocketDebuggerUrl']:
                                url = tab.get('url', '')
                                if url and (url.startswith('http://') or url.startswith('https://')):
                                    if not url.startswith('chrome://') and not url.startswith('about:'):
                                        return url
                    
                except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError):
                    # Skip this port and try next one
                    continue
                except:
                    continue
        except:
            pass
        return None
    
    def _get_url_from_window_title_simple(self):
        """Simple window title URL detection - ONLY checks the currently active/foreground browser window"""
        try:
            if sys.platform == "win32":
                import win32gui
                import win32con
                import re
                
                browser_keywords = ['chrome', 'edge', 'firefox', 'opera', 'brave', 'vivaldi', 'microsoft edge']
                url_pattern = r'https?://[^\s<>"\'\)]+'
                
                # STRICT: Only check the foreground/active window
                try:
                    hwnd = win32gui.GetForegroundWindow()
                    if not hwnd:
                        return None
                    
                    window_title = win32gui.GetWindowText(hwnd)
                    
                    if not window_title:
                        return None
                    
                    title_lower = window_title.lower()
                    
                    # Check if it's a browser window - verify it's actually a browser
                    is_browser = False
                    for keyword in browser_keywords:
                        if keyword in title_lower:
                            # Exclude extension windows and popups
                            if 'extension' not in title_lower and 'popup' not in title_lower:
                                is_browser = True
                                break
                    
                    if not is_browser:
                        return None
                    
                    # Check if window is minimized and try to restore it
                    try:
                        placement = win32gui.GetWindowPlacement(hwnd)
                        if placement[1] == win32con.SW_SHOWMINIMIZED:
                            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
                            import time
                            time.sleep(0.1)  # Brief wait for window to restore
                            window_title = win32gui.GetWindowText(hwnd)  # Refresh title
                    except:
                        pass
                    
                    # Extract URL from the active window title
                    match = re.search(url_pattern, window_title)
                    if match:
                        url = match.group(0).rstrip('.,;:!?)')
                        if url.startswith('http://') or url.startswith('https://'):
                            domain_part = url.split('://')[1].split('/')[0]
                            if '.' in domain_part and len(domain_part) > 3:
                                # Double-check: verify the window is still active (prevent race conditions)
                                current_hwnd = win32gui.GetForegroundWindow()
                                if current_hwnd == hwnd:
                                    return url
                except Exception as e:
                    pass
        except Exception as e:
            pass
        return None
    
    def _try_get_url_via_keyboard(self):
        """Try to get URL by using keyboard automation - works on currently active browser window"""
        try:
            if not PYNPUT_AVAILABLE:
                return None
            
            import time
            import win32gui
            from pynput.keyboard import Key, Controller as KeyboardController
            
            # Get the active window
            active_hwnd = win32gui.GetForegroundWindow()
            if not active_hwnd:
                return None
            
            active_window_title = win32gui.GetWindowText(active_hwnd)
            if not active_window_title:
                return None
            
            # Verify it's a browser window
            title_lower = active_window_title.lower()
            browser_keywords = ['chrome', 'edge', 'firefox', 'opera', 'brave', 'vivaldi', 'microsoft edge']
            is_browser = False
            for keyword in browser_keywords:
                if keyword in title_lower:
                    # Exclude extension windows and popups
                    if 'extension' not in title_lower and 'popup' not in title_lower:
                        is_browser = True
                        break
            
            if not is_browser:
                # If active window is not a browser, try to find and activate one
                browser_hwnd = self._find_and_activate_any_browser_window()
                if browser_hwnd:
                    active_hwnd = browser_hwnd
                    active_window_title = win32gui.GetWindowText(active_hwnd)
                    time.sleep(0.2)  # Wait for window to activate
                else:
                    return None  # No browser window found
            
            keyboard = KeyboardController()
            
            # Ensure browser window is in foreground (with timeout protection)
            try:
                win32gui.SetForegroundWindow(active_hwnd)
                time.sleep(0.1)  # Reduced wait time
            except:
                # If SetForegroundWindow fails, try alternative
                try:
                    import win32con
                    win32gui.ShowWindow(active_hwnd, win32con.SW_RESTORE)
                    win32gui.ShowWindow(active_hwnd, win32con.SW_SHOW)
                    time.sleep(0.1)
                except:
                    pass
            
            # Don't verify window is active - just proceed (prevents hangs)
            # The keyboard automation will work even if window isn't perfectly focused
            
            # Press Ctrl+L to focus address bar
            keyboard.press(Key.ctrl)
            keyboard.press('l')
            keyboard.release('l')
            keyboard.release(Key.ctrl)
            
            time.sleep(0.3)  # Wait for address bar to focus
            
            # Press Ctrl+A to select all
            keyboard.press(Key.ctrl)
            keyboard.press('a')
            keyboard.release('a')
            keyboard.release(Key.ctrl)
            
            time.sleep(0.2)  # Wait for selection
            
            # Press Ctrl+C to copy URL
            keyboard.press(Key.ctrl)
            keyboard.press('c')
            keyboard.release('c')
            keyboard.release(Key.ctrl)
            
            time.sleep(0.3)  # Wait for clipboard to update
            
            # Try to get URL from clipboard (with retries)
            for attempt in range(3):  # Increased attempts for reliability
                try:
                    clipboard_url = self.root.clipboard_get()
                    if clipboard_url:
                        clipboard_url = clipboard_url.strip()
                        if clipboard_url.startswith('http://') or clipboard_url.startswith('https://'):
                            # Validate it's a proper URL
                            try:
                                url_parts = clipboard_url.split('://')
                                if len(url_parts) == 2:
                                    domain_part = url_parts[1].split('/')[0]
                                    if '.' in domain_part and len(domain_part) > 3:
                                        return clipboard_url
                            except:
                                pass
                    if attempt < 2:
                        time.sleep(0.2)  # Wait between attempts
                except:
                    if attempt < 2:
                        time.sleep(0.2)
        except:
            pass
        return None
    
    def _find_and_activate_browser_window(self):
        """Find and activate the currently active browser window - STRICT: only active window"""
        if sys.platform == "win32":
            try:
                import win32gui
                import win32con
                
                browser_keywords = ['chrome', 'edge', 'firefox', 'opera', 'brave', 'vivaldi', 'microsoft edge']
                
                # STRICT: Only check the foreground/active window
                active_hwnd = win32gui.GetForegroundWindow()
                if not active_hwnd:
                    return None
                
                active_window_title = win32gui.GetWindowText(active_hwnd)
                if not active_window_title:
                    return None
                
                title_lower = active_window_title.lower()
                
                # Check if active window is a browser
                if any(keyword in title_lower for keyword in browser_keywords):
                    # Exclude extension windows and popups
                    if 'extension' not in title_lower and 'popup' not in title_lower:
                        # Try to ensure window is not minimized
                        try:
                            placement = win32gui.GetWindowPlacement(active_hwnd)
                            if placement[1] == win32con.SW_SHOWMINIMIZED:
                                win32gui.ShowWindow(active_hwnd, win32con.SW_RESTORE)
                                import time
                                time.sleep(0.1)
                            win32gui.SetForegroundWindow(active_hwnd)
                            return active_hwnd
                        except:
                            return active_hwnd
            except Exception as e:
                pass
        return None
    
    def _find_and_activate_any_browser_window(self):
        """Find and activate ANY browser window (not just active one) - used when tool window is active"""
        if sys.platform == "win32":
            try:
                import win32gui
                import win32con
                
                browser_keywords = ['chrome', 'edge', 'firefox', 'opera', 'brave', 'vivaldi', 'microsoft edge']
                browser_windows = []
                
                def enum_windows_callback(hwnd, param):
                    try:
                        if win32gui.IsWindowVisible(hwnd):
                            window_title = win32gui.GetWindowText(hwnd)
                            if window_title:
                                title_lower = window_title.lower()
                                if any(keyword in title_lower for keyword in browser_keywords):
                                    # Exclude extension windows and popups
                                    if 'extension' not in title_lower and 'popup' not in title_lower:
                                        param.append((hwnd, window_title))
                    except:
                        pass
                    return True
                
                win32gui.EnumWindows(enum_windows_callback, browser_windows)
                
                if browser_windows:
                    # Use the first browser window found
                    browser_hwnd, window_title = browser_windows[0]
                    
                    # Try to bring window to foreground (with error handling to prevent hangs)
                    try:
                        placement = win32gui.GetWindowPlacement(browser_hwnd)
                        if placement[1] == win32con.SW_SHOWMINIMIZED:
                            win32gui.ShowWindow(browser_hwnd, win32con.SW_RESTORE)
                            import time
                            time.sleep(0.05)  # Reduced wait time
                        
                        # Try to set foreground window (may fail on some systems, that's OK)
                        try:
                            win32gui.SetForegroundWindow(browser_hwnd)
                        except:
                            # If SetForegroundWindow fails, try alternative method
                            try:
                                win32gui.ShowWindow(browser_hwnd, win32con.SW_RESTORE)
                                win32gui.ShowWindow(browser_hwnd, win32con.SW_SHOW)
                            except:
                                pass
                        
                        return browser_hwnd
                    except Exception as e:
                        # Return hwnd anyway - we can still try to detect
                        return browser_hwnd
            except Exception as e:
                pass  # Silently fail to prevent error spam
        return None
    
    def _show_detection_failed_dialog(self):
        """Show dialog when URL detection fails with options to try again, paste, or enter manually"""
        dialog = tk.Toplevel(self.root)
        dialog.title("URL Not Detected")
        dialog.geometry("500x350")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (350 // 2)
        dialog.geometry(f"500x350+{x}+{y}")
        
        # Main message frame
        message_frame = ttk.Frame(dialog, padding="20")
        message_frame.pack(fill=tk.BOTH, expand=True)
        
        # Icon and title
        ttk.Label(message_frame, 
                 text="⚠️ URL Not Detected",
                 font=("Arial", 12, "bold")).pack(pady=(0, 10))
        
        ttk.Label(message_frame, 
                 text="Could not automatically detect the URL from your opened browser.",
                 font=("Arial", 9),
                 wraplength=450,
                 justify=tk.CENTER).pack(pady=(0, 15))
        
        # Possible reasons
        reasons_frame = ttk.LabelFrame(message_frame, text="Possible reasons:", padding="10")
        reasons_frame.pack(pady=(0, 15), padx=10, fill=tk.X)
        
        reasons = [
            "• Browser window might be minimized or not visible",
            "• Browser not started with remote debugging (port 9222)",
            "• URL might not be visible in window title",
            "• Multiple browser windows open (tool may detect wrong one)"
        ]
        for reason in reasons:
            ttk.Label(reasons_frame, text=reason, font=("Arial", 8)).pack(anchor=tk.W, pady=2)
        
        # Buttons frame
        buttons_frame = ttk.Frame(message_frame)
        buttons_frame.pack(pady=10, fill=tk.X)
        
        # Instructions for manual method
        instruction_frame = ttk.LabelFrame(message_frame, text="📋 Easiest Method:", padding="10")
        instruction_frame.pack(pady=(0, 15), padx=10, fill=tk.X)
        
        steps = [
            "1. Click in your browser's address bar (or press Ctrl+L)",
            "2. Press Ctrl+A to select all",
            "3. Press Ctrl+C to copy the URL",
            "4. Click 'Paste URL from Clipboard' button below"
        ]
        for step in steps:
            ttk.Label(instruction_frame, text=step, font=("Arial", 8)).pack(anchor=tk.W, pady=2)
        
        # Button 1: Paste from Clipboard (make this the primary option)
        btn_paste = ttk.Button(buttons_frame, 
                              text="📋 Paste URL from Clipboard (Recommended)",
                              command=lambda: self._handle_detection_option(dialog, "paste"),
                              width=40)
        btn_paste.pack(pady=5, padx=20, fill=tk.X)
        
        # Button 2: Try Again (Detect from Browser)
        btn_try_again = ttk.Button(buttons_frame, 
                                   text="🔄 Try Again - Auto Detect from Browser",
                                   command=lambda: self._retry_detection(dialog),
                                   width=40)
        btn_try_again.pack(pady=5, padx=20, fill=tk.X)
        
        # Add instruction label for Try Again
        ttk.Label(buttons_frame, 
                 text="💡 For 'Try Again': Make sure browser window is active/focused",
                 font=("Arial", 8),
                 foreground="blue",
                 wraplength=400,
                 justify=tk.CENTER).pack(pady=(0, 5))
        
        # Button 3: Enter Manually
        btn_manual = ttk.Button(buttons_frame, 
                                text="✏️ Enter URL Manually",
                                command=lambda: self._handle_detection_option(dialog, "manual"),
                                width=40)
        btn_manual.pack(pady=5, padx=20, fill=tk.X)
        
        # Cancel button
        ttk.Button(message_frame, 
                  text="Cancel",
                  command=dialog.destroy).pack(pady=(10, 0))
    
    def _retry_detection(self, dialog):
        """Retry URL detection from browser"""
        dialog.destroy()
        # Show message to user with countdown
        response = messagebox.askokcancel("Retrying Detection", 
                          "Please make sure:\n\n"
                          "1. Your browser window is open and visible\n"
                          "2. The browser window is ACTIVE/FOCUSED (click on it)\n"
                          "3. You are on the page with the URL you want to detect\n\n"
                          "Click OK to start detection (will try to activate browser window automatically)...")
        if response:
            # Call the detection function again
            self.detect_url_from_browser()
    
    def _handle_detection_option(self, dialog, option):
        """Handle option selected from detection failed dialog"""
        dialog.destroy()
        if option == "paste":
                self._paste_url_from_clipboard()
        elif option == "manual":
            self.manual_url_entry.focus()
            self.manual_url_entry.select_range(0, tk.END)
            messagebox.showinfo("Manual URL Entry", 
                              "Please enter the URL in the 'Manual Override' section below,\n"
                              "then click 'Set URL' button.")
    
    def _try_detect_from_active_browser_window(self):
        """Try to detect URL from the currently active/focused browser window"""
        if sys.platform == "win32":
            try:
                import win32gui
                import re
                
                # Get the foreground (active) window
                hwnd = win32gui.GetForegroundWindow()
                window_title = win32gui.GetWindowText(hwnd)
                
                if window_title:
                    # Check if it's a browser window
                    title_lower = window_title.lower()
                    browser_keywords = ['chrome', 'edge', 'microsoft edge', 'firefox', 'opera', 'brave', 'vivaldi']
                    
                    if any(browser in title_lower for browser in browser_keywords):
                        # Try to extract URL from title
                        url_pattern = r'https?://[^\s<>"\'\)]+'
                        match = re.search(url_pattern, window_title)
                        if match:
                            url = match.group(0).rstrip('.,;:!?)')
                            if url.startswith('http://') or url.startswith('https://'):
                                # Validate domain
                                if '.' in url.split('://')[1].split('/')[0]:
                                    return url
                
                # Also check all browser windows and return the first URL found
                windows = []
                def enum_windows_callback(hwnd, param):
                    try:
                        if win32gui.IsWindowVisible(hwnd):
                            window_title = win32gui.GetWindowText(hwnd)
                            if window_title:
                                title_lower = window_title.lower()
                                if any(browser in title_lower for browser in browser_keywords):
                                    param.append(window_title)
                    except:
                        pass
                    return True
                
                win32gui.EnumWindows(enum_windows_callback, windows)
                
                # Try to extract URL from any browser window
                url_pattern = r'https?://[^\s<>"\'\)]+'
                for window_title in windows:
                    match = re.search(url_pattern, window_title)
                    if match:
                        url = match.group(0).rstrip('.,;:!?)')
                        if url.startswith('http://') or url.startswith('https://'):
                            if '.' in url.split('://')[1].split('/')[0]:
                                return url
            except Exception as e:
                self.log_message(f"Error detecting from active window: {e}", "WARNING")
        return None
    
    def _check_if_browser_is_open(self):
        """Check if any browser window is open"""
        if sys.platform == "win32":
            try:
                import win32gui
                def enum_windows_callback(hwnd, browsers):
                    window_title = win32gui.GetWindowText(hwnd)
                    if window_title:
                        title_lower = window_title.lower()
                        if any(browser in title_lower for browser in ['chrome', 'edge', 'microsoft edge', 'firefox', 'opera', 'brave']):
                            browsers.append(hwnd)
                    return True
                
                browsers = []
                win32gui.EnumWindows(enum_windows_callback, browsers)
                return len(browsers) > 0
            except:
                return False
        return False
    
    def _handle_url_change_with_confirmation(self, new_url, browser=None, mode=None):
        """Handle URL change with user confirmation"""
        # Format mode and browser names for display
        if browser and mode:
            mode_display = "Normal Mode" if mode == "normal" else "Incognito Mode" if browser == "chrome" else "New Private Window"
            browser_display = {"chrome": "Google Chrome", "edge": "Microsoft Edge", "firefox": "Mozilla Firefox"}.get(browser, browser.capitalize())
        else:
            mode_display = None
            browser_display = None
        
        if self.current_url:
            # URL already set - ask user if they want to update
            message = f"A different URL was detected in your browser:\n\n"
            message += f"New URL: {new_url}\n"
            if browser_display and mode_display:
                message += f"Browser: {browser_display}\n"
                message += f"Mode: {mode_display}\n"
            message += f"\nCurrent URL: {self.current_url}\n\n"
            message += f"What do you want to do?\n\n"
            message += f"• Click 'Yes' to update to the new URL\n"
            message += f"• Click 'No' to continue with the existing URL"
            
            response = messagebox.askyesno("URL Changed Detected", message)
            if response:
                # User wants to update to new URL
                try:
                    self.browser_monitor._handle_url_change(new_url)
                    update_message = f"URL successfully updated!\n\n"
                    update_message += f"New URL: {new_url}\n"
                    if browser_display and mode_display:
                        update_message += f"Browser: {browser_display}\n"
                        update_message += f"Mode: {mode_display}\n"
                    update_message += f"Module: {self.current_module}\n"
                    update_message += f"Page: {self.current_page}\n\n"
                    update_message += f"You can now continue capturing test cases."
                    messagebox.showinfo("URL Updated", update_message)
                except Exception as e:
                    self.log_message(f"❌ Error updating URL: {e}", "ERROR")
                    messagebox.showerror("Error", f"Error updating URL:\n{e}")
            else:
                # User wants to keep existing URL
                self.log_message(f"User chose to keep existing URL: {self.current_url}", "INFO")
        else:
            # No URL set yet - just update it
            try:
                self.browser_monitor._handle_url_change(new_url)
                detect_message = f"URL successfully detected and updated!\n\n"
                detect_message += f"URL: {new_url}\n"
                if browser_display and mode_display:
                    detect_message += f"Browser: {browser_display}\n"
                    detect_message += f"Mode: {mode_display}\n"
                detect_message += f"Module: {self.current_module}\n"
                detect_message += f"Page: {self.current_page}\n\n"
                detect_message += f"You can now start capturing test cases."
                messagebox.showinfo("URL Detected", detect_message)
            except Exception as e:
                self.log_message(f"❌ Error updating URL: {e}", "ERROR")
                messagebox.showerror("Error", f"Error updating URL:\n{e}")
    
    def _show_url_detection_dialog(self):
        """Show dialog with options: Update from Browser, Paste from Clipboard, or Enter Manually"""
        dialog = tk.Toplevel(self.root)
        dialog.title("URL Not Auto-Detected")
        dialog.geometry("550x400")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()  # Make it modal
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (550 // 2)
        y = (dialog.winfo_screenheight() // 2) - (400 // 2)
        dialog.geometry(f"550x400+{x}+{y}")
        
        # Main message
        message_frame = ttk.Frame(dialog, padding="20")
        message_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(message_frame, 
                 text="Could not automatically detect the current browser URL.",
                 font=("Arial", 10, "bold")).pack(pady=(0, 10))
        
        ttk.Label(message_frame, 
                 text="This is normal if you opened the browser manually.",
                 font=("Arial", 9)).pack(pady=(0, 10))
        
        # Quick instructions
        quick_instructions = ttk.Label(message_frame, 
                                      text="⚡ QUICKEST WAY:\n1. Copy URL from browser (Ctrl+L, Ctrl+C)\n2. Click 'Paste URL Manually' button",
                                      font=("Arial", 8, "bold"),
                                      foreground="green",
                                      justify=tk.CENTER)
        quick_instructions.pack(pady=(0, 15))
        
        ttk.Label(message_frame, 
                 text="Or choose an option below:",
                 font=("Arial", 9, "bold")).pack(pady=(0, 10))
        
        ttk.Label(message_frame, 
                 text="💡 Tip: Make sure your browser is open and navigate to the desired URL,\nthen try 'Detect URL from Browser' again.",
                 font=("Arial", 8),
                 foreground="blue",
                 justify=tk.CENTER).pack(pady=(0, 15))
        
        ttk.Label(message_frame, 
                 text="Note: Make sure your browser is open and navigate to the desired URL,\nthen try 'Detect URL from Browser' again.",
                 font=("Arial", 8),
                 foreground="blue",
                 justify=tk.CENTER).pack(pady=(0, 10))
        
        # Buttons frame
        buttons_frame = ttk.Frame(message_frame)
        buttons_frame.pack(pady=10)
        
        # Button 1: Paste from Clipboard
        btn_clipboard = ttk.Button(buttons_frame, 
                                  text="1. Paste URL from Clipboard",
                                  command=lambda: self._handle_browser_option(dialog, "clipboard"),
                                  width=40)
        btn_clipboard.pack(pady=8, padx=20, fill=tk.X)
        
        # Button 2: Enter Manually
        btn_manual = ttk.Button(buttons_frame, 
                               text="2. Enter URL Manually",
                               command=lambda: self._handle_browser_option(dialog, "manual"),
                               width=40)
        btn_manual.pack(pady=8, padx=20, fill=tk.X)
        
        # Instructions
        instructions = ttk.Label(message_frame, 
                                text="💡 Tip: The 'Paste URL Manually' button is the fastest method!\nJust copy URL from browser (Ctrl+L, Ctrl+C) and click it.",
                                font=("Arial", 8), 
                                foreground="blue",
                                justify=tk.CENTER)
        instructions.pack(pady=(15, 0))
        
        # Cancel button
        ttk.Button(message_frame, 
                  text="Cancel",
                  command=dialog.destroy).pack(pady=(15, 0))
    
    def _handle_browser_option(self, dialog, option):
        """Handle the selected option"""
        dialog.destroy()
        
        if option == "clipboard":
            # Try to get URL from clipboard
            self._paste_url_from_clipboard()
        elif option == "manual":
            # Focus on manual URL entry
            self.manual_url_entry.focus()
            self.manual_url_entry.select_range(0, tk.END)
            self.log_message("Waiting for manual URL entry...", "INFO")
            messagebox.showinfo("Enter URL", 
                              "Please enter the current URL in the 'Manual Override' section below,\n"
                              "then click 'Set URL' button.")
    
    def _paste_url_from_clipboard(self):
        """Paste URL from clipboard and set it"""
        try:
            clipboard_text = self.root.clipboard_get()
            if clipboard_text:
                clipboard_text = clipboard_text.strip()
                # Check if it looks like a URL
                if clipboard_text.startswith('http://') or clipboard_text.startswith('https://'):
                    # Validate URL format
                    if '.' in clipboard_text.split('://')[1].split('/')[0]:
                        # Set the URL
                        self.manual_url_var.set(clipboard_text)
                        self.log_message(f"✅ URL pasted from clipboard: {clipboard_text}", "SUCCESS")
                        # Automatically set the URL
                        self.set_manual_url()
                        messagebox.showinfo("URL Pasted", 
                                          f"URL pasted from clipboard and set!\n\n"
                                          f"URL: {clipboard_text}\n\n"
                                          f"Module and Page will be auto-detected.")
                        return
                    else:
                        messagebox.showerror("Invalid URL", 
                                           f"The clipboard doesn't contain a valid URL.\n\n"
                                           f"Clipboard content: {clipboard_text[:50]}...\n\n"
                                           f"Please copy a valid URL from your browser address bar.")
                else:
                    messagebox.showerror("Invalid URL", 
                                       f"The clipboard doesn't contain a URL.\n\n"
                                       f"Clipboard content: {clipboard_text[:50]}...\n\n"
                                       f"Please:\n"
                                       f"1. Click in your browser address bar (Ctrl+L)\n"
                                       f"2. Copy the URL (Ctrl+C)\n"
                                       f"3. Try again")
            else:
                messagebox.showwarning("Clipboard Empty", 
                                     "The clipboard is empty.\n\n"
                                     "Please:\n"
                                     "1. Open your browser\n"
                                     "2. Click in the address bar (Ctrl+L)\n"
                                     "3. Copy the URL (Ctrl+C)\n"
                                     "4. Try again")
        except tk.TclError:
            messagebox.showwarning("Clipboard Error", 
                                 "Could not read from clipboard.\n\n"
                                 "Please manually enter the URL in the 'Manual Override' section.")
        except Exception as e:
            self.log_message(f"Error reading clipboard: {e}", "ERROR")
            messagebox.showerror("Error", f"Error reading clipboard: {e}")
    
    def clear_url(self):
        """Clear the current URL and reset browser state"""
        try:
            response = messagebox.askyesno("Clear URL", 
                                          "Are you sure you want to clear the current URL?\n\n"
                                          "This will reset:\n"
                                          "- Current URL\n"
                                          "- Module and Page detection\n"
                                          "- Browser state\n\n"
                                          "You can set a new URL using 'Detect URL from Browser' or 'Paste URL Manually'.")
            if response:
                # FORCE RESET: Reset detection state (important: allows detection to work again after clearing)
                self._detection_in_progress = False
                self._url_detected_flag = False  # Reset detection flag
                if hasattr(self, '_detection_start_time'):
                    delattr(self, '_detection_start_time')
                
                # Hide loading indicator if showing
                self._hide_loading_indicator()
                
                # Force reset detection button state (check actual state first)
                if hasattr(self, 'detect_url_button'):
                    try:
                        current_state = self.detect_url_button.cget('state')
                        current_text = self.detect_url_button.cget('text')
                        if current_state == 'disabled' or 'Detecting' in current_text:
                            self.log_message("🔧 Resetting stuck button state...", "INFO")
                    except Exception as e:
                        self.log_message(f"Error checking button state: {e}", "WARNING")
                
                self._reset_detection_button()
                
                # Log state reset for debugging
                self.log_message(f"🔧 Detection state reset: _detection_in_progress={self._detection_in_progress}", "INFO")
                
                # Set flag to prevent browser monitor from re-setting URL immediately
                import time
                self._url_cleared_flag = True
                self._url_clear_time = time.time()
                
                # Clear URL and related data
                self.current_url = ""
                self.current_module = ""
                self.current_page = ""
                self.current_tab = ""
                
                # Clear browser monitor state
                try:
                    self.browser_monitor.current_url = ""
                    self.browser_monitor.current_module = ""
                    self.browser_monitor.current_page = ""
                except Exception as e:
                    self.log_message(f"Error clearing browser monitor: {e}", "WARNING")
                
                # Reset the flag after 3 seconds (allows manual detection to work)
                def reset_clear_flag():
                    try:
                        self._url_cleared_flag = False
                        if hasattr(self, '_url_clear_time'):
                            delattr(self, '_url_clear_time')
                    except Exception:
                        pass  # Silently handle errors
                
                try:
                    if hasattr(self, 'root') and self.root:
                        self.root.after(3000, reset_clear_flag)  # Reset flag after 3 seconds
                except (tk.TclError, AttributeError) as e:
                    self.log_message(f"Warning: Could not schedule flag reset: {e}", "WARNING")
                
                # Clear manual URL entry
                try:
                    if hasattr(self, 'manual_url_var'):
                        self.manual_url_var.set("")
                except Exception as e:
                    self.log_message(f"Error clearing manual URL entry: {e}", "WARNING")
                
                # Update UI labels
                try:
                    if hasattr(self, 'url_label'):
                        self.url_label.config(text="URL: Not detected")
                    if hasattr(self, 'session_info_label'):
                        self.session_info_label.config(text="Module: Auto-detected | Page: Auto-detected")
                    if hasattr(self, 'url_status_label'):
                        self.url_status_label.config(text="⚠ URL cleared. Use 'Detect URL from Browser' or 'Paste URL Manually' to set a new URL", 
                                                   foreground="orange")
                except Exception as e:
                    self.log_message(f"Error updating UI labels: {e}", "WARNING")
                
                # Log the action
                self.log_message("✅ URL cleared successfully", "SUCCESS")
                self.log_message("✅ Ready for new detection. Click 'Detect URL from Browser' to start.", "INFO")
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            self.log_message(f"❌ Error clearing URL: {e}", "ERROR")
            self.log_message(f"Error details: {error_details}", "ERROR")
            messagebox.showerror("Error", f"An error occurred while clearing URL:\n{e}\n\nPlease try again.")
    
    def _show_browser_selection_dialog(self):
        """Show browser selection dialog"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Select Browser")
        dialog.geometry("450x400")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (dialog.winfo_screenheight() // 2) - (400 // 2)
        dialog.geometry(f"450x400+{x}+{y}")
        
        # Main message
        message_frame = ttk.Frame(dialog, padding="20")
        message_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(message_frame, 
                 text="Select Browser",
                 font=("Arial", 12, "bold")).pack(pady=(0, 10))
        
        ttk.Label(message_frame, 
                 text="Which browser would you like to use?",
                 font=("Arial", 9)).pack(pady=(0, 20))
        
        # Browser buttons frame
        browsers_frame = ttk.Frame(message_frame)
        browsers_frame.pack(pady=10, fill=tk.BOTH, expand=True)
        
        selected_browser = [None]  # Use list to allow modification in nested function
        
        browsers = [
            ("Google Chrome", "chrome"),
            ("Microsoft Edge", "edge"),
            ("Mozilla Firefox", "firefox"),
            ("Opera", "opera"),
            ("Brave", "brave"),
            ("Vivaldi", "vivaldi")
        ]
        
        for browser_name, browser_id in browsers:
            btn = ttk.Button(browsers_frame, 
                            text=browser_name,
                            command=lambda b=browser_id: self._select_browser(dialog, b, selected_browser),
                            width=35)
            btn.pack(pady=8, padx=20, fill=tk.X)
        
        # Cancel button
        ttk.Button(message_frame, 
                  text="Cancel",
                  command=dialog.destroy).pack(pady=(15, 0))
        
        dialog.wait_window()
        return selected_browser[0]
    
    def _select_browser(self, dialog, browser_id, selected_browser):
        """Handle browser selection"""
        selected_browser[0] = browser_id
        dialog.destroy()
    
    def _show_mode_selection_dialog(self, browser):
        """Show mode selection dialog (normal or incognito)"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Select Mode")
        dialog.geometry("400x250")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (dialog.winfo_screenheight() // 2) - (250 // 2)
        dialog.geometry(f"400x250+{x}+{y}")
        
        # Main message
        message_frame = ttk.Frame(dialog, padding="20")
        message_frame.pack(fill=tk.BOTH, expand=True)
        
        browser_names = {
            "chrome": "Google Chrome",
            "edge": "Microsoft Edge",
            "firefox": "Mozilla Firefox",
            "opera": "Opera",
            "brave": "Brave",
            "vivaldi": "Vivaldi"
        }
        
        browser_name = browser_names.get(browser, browser.capitalize())
        
        ttk.Label(message_frame, 
                 text=f"Open {browser_name} in:",
                 font=("Arial", 11, "bold")).pack(pady=(0, 20))
        
        # Mode buttons frame
        modes_frame = ttk.Frame(message_frame)
        modes_frame.pack(pady=10, fill=tk.BOTH, expand=True)
        
        selected_mode = [None]
        
        # Normal mode button
        btn_normal = ttk.Button(modes_frame, 
                               text="Normal Mode",
                               command=lambda: self._select_mode(dialog, "normal", selected_mode),
                               width=30)
        btn_normal.pack(pady=10, padx=20, fill=tk.X)
        
        # Incognito/Private mode button
        mode_name = "Incognito Mode" if browser in ["chrome", "edge", "opera", "brave", "vivaldi"] else "Private Mode"
        btn_incognito = ttk.Button(modes_frame, 
                                  text=mode_name,
                                  command=lambda: self._select_mode(dialog, "incognito", selected_mode),
                                  width=30)
        btn_incognito.pack(pady=10, padx=20, fill=tk.X)
        
        # Cancel button
        ttk.Button(message_frame, 
                  text="Cancel",
                  command=dialog.destroy).pack(pady=(15, 0))
        
        dialog.wait_window()
        return selected_mode[0]
    
    def _select_mode(self, dialog, mode, selected_mode):
        """Handle mode selection"""
        selected_mode[0] = mode
        dialog.destroy()
    
    def _show_url_input_dialog(self):
        """Show dialog to ask user for URL to open in browser"""
        url = simpledialog.askstring("Enter URL", 
                                    "Enter the URL to open in the browser:\n\n"
                                    "(Leave empty to open browser's default page)\n"
                                    "(You can cancel and URL will be detected automatically)",
                                    initialvalue="")
        return url if url else None
    
    def _launch_browser(self, browser, mode, url=None):
        """Launch browser in specified mode with remote debugging"""
        self.log_message(f"Launching {browser} in {mode} mode...", "INFO")
        
        # Use provided URL or default to empty (browser will open to default page)
        base_url = url if url else ""
        port = 9222
        
        try:
            if browser == "chrome":
                # Create a unique user data directory for this session
                user_data_dir = os.path.join(tempfile.gettempdir(), f"chrome_debug_profile_{browser}_{mode}")
                
                if mode == "incognito":
                    cmd = [
                        "chrome.exe" if os.name == 'nt' else "google-chrome",
                        f"--remote-debugging-port={port}",
                        f"--user-data-dir={user_data_dir}",
                        "--incognito"
                    ]
                    if base_url:
                        cmd.append(base_url)
                else:
                    cmd = [
                        "chrome.exe" if os.name == 'nt' else "google-chrome",
                        f"--remote-debugging-port={port}",
                        f"--user-data-dir={user_data_dir}"
                    ]
                    if base_url:
                        cmd.append(base_url)
                # Try common Chrome paths on Windows
                if os.name == 'nt':
                    chrome_paths = [
                        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                        os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe")
                    ]
                    for path in chrome_paths:
                        if os.path.exists(path):
                            cmd[0] = path
                            break
                    else:
                        cmd[0] = "chrome"  # Fallback to PATH
                        
            elif browser == "edge":
                # Create a unique user data directory for this session
                user_data_dir = os.path.join(tempfile.gettempdir(), f"edge_debug_profile_{browser}_{mode}")
                
                if mode == "incognito":
                    cmd = [
                        "msedge.exe" if os.name == 'nt' else "microsoft-edge",
                        f"--remote-debugging-port={port}",
                        f"--user-data-dir={user_data_dir}",
                        "--inprivate"
                    ]
                    if base_url:
                        cmd.append(base_url)
                else:
                    cmd = [
                        "msedge.exe" if os.name == 'nt' else "microsoft-edge",
                        f"--remote-debugging-port={port}",
                        f"--user-data-dir={user_data_dir}"
                    ]
                    if base_url:
                        cmd.append(base_url)
                # Try common Edge paths on Windows
                if os.name == 'nt':
                    edge_paths = [
                        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
                        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"
                    ]
                    for path in edge_paths:
                        if os.path.exists(path):
                            cmd[0] = path
                            break
                    else:
                        cmd[0] = "msedge"  # Fallback to PATH
                        
            elif browser == "firefox":
                if mode == "incognito":
                    cmd = [
                        "firefox.exe" if os.name == 'nt' else "firefox",
                        "-private-window"
                    ]
                    if base_url:
                        cmd.append(base_url)
                else:
                    cmd = [
                        "firefox.exe" if os.name == 'nt' else "firefox"
                    ]
                    if base_url:
                        cmd.append(base_url)
                # Firefox doesn't support remote debugging the same way
                self.log_message("⚠️ Firefox doesn't support remote debugging like Chrome/Edge", "WARNING")
                self.log_message("Please use Chrome or Edge for automatic URL detection, or use Manual Override", "INFO")
                
            elif browser == "opera":
                # Create a unique user data directory for this session
                user_data_dir = os.path.join(tempfile.gettempdir(), f"opera_debug_profile_{browser}_{mode}")
                
                if mode == "incognito":
                    cmd = [
                        "opera.exe" if os.name == 'nt' else "opera",
                        f"--remote-debugging-port={port}",
                        f"--user-data-dir={user_data_dir}",
                        "--private"
                    ]
                    if base_url:
                        cmd.append(base_url)
                else:
                    cmd = [
                        "opera.exe" if os.name == 'nt' else "opera",
                        f"--remote-debugging-port={port}",
                        f"--user-data-dir={user_data_dir}"
                    ]
                    if base_url:
                        cmd.append(base_url)
                    
            elif browser == "brave":
                # Create a unique user data directory for this session
                user_data_dir = os.path.join(tempfile.gettempdir(), f"brave_debug_profile_{browser}_{mode}")
                
                if mode == "incognito":
                    cmd = [
                        "brave.exe" if os.name == 'nt' else "brave-browser",
                        f"--remote-debugging-port={port}",
                        f"--user-data-dir={user_data_dir}",
                        "--incognito"
                    ]
                    if base_url:
                        cmd.append(base_url)
                else:
                    cmd = [
                        "brave.exe" if os.name == 'nt' else "brave-browser",
                        f"--remote-debugging-port={port}",
                        f"--user-data-dir={user_data_dir}"
                    ]
                    if base_url:
                        cmd.append(base_url)
                    
            elif browser == "vivaldi":
                # Create a unique user data directory for this session
                user_data_dir = os.path.join(tempfile.gettempdir(), f"vivaldi_debug_profile_{browser}_{mode}")
                
                if mode == "incognito":
                    cmd = [
                        "vivaldi.exe" if os.name == 'nt' else "vivaldi",
                        f"--remote-debugging-port={port}",
                        f"--user-data-dir={user_data_dir}",
                        "--incognito"
                    ]
                    if base_url:
                        cmd.append(base_url)
                else:
                    cmd = [
                        "vivaldi.exe" if os.name == 'nt' else "vivaldi",
                        f"--remote-debugging-port={port}",
                        f"--user-data-dir={user_data_dir}"
                    ]
                    if base_url:
                        cmd.append(base_url)
            else:
                self.log_message(f"Unknown browser: {browser}", "ERROR")
                return
            
            # Launch browser
            subprocess.Popen(cmd, shell=False)
            self.log_message(f"✅ {browser.capitalize()} launched in {mode} mode with remote debugging on port {port}", "SUCCESS")
            if base_url:
                self.log_message(f"Browser will open: {base_url}", "INFO")
            else:
                self.log_message("Browser will open to default page (no URL specified)", "INFO")
            
            # Start browser monitoring if not already started
            if not self.browser_monitor.monitoring:
                self.log_message("Starting browser monitoring...", "INFO")
                if self.browser_monitor.start_monitoring():
                    self.log_message("✅ Browser monitoring started", "SUCCESS")
                else:
                    self.log_message("⚠️ Failed to start browser monitoring", "WARNING")
            
            self.log_message("Waiting 5 seconds for browser to fully start, then attempting URL detection...", "INFO")
            
            # Wait a bit for browser to start, then try to detect URL (retry multiple times)
            self.root.after(5000, lambda: self._try_detect_url_after_launch(attempt=1))
            
        except Exception as e:
            self.log_message(f"❌ Error launching browser: {e}", "ERROR")
            messagebox.showerror("Error", f"Failed to launch {browser}:\n{e}\n\nPlease launch the browser manually and use 'Detect URL from Browser' or Manual Override.")
    
    def _try_detect_url_after_launch(self, attempt=1, max_attempts=5):
        """Try to detect URL after browser launch (with retries)"""
        try:
            self.log_message(f"Attempting to detect URL from launched browser (attempt {attempt}/{max_attempts})...", "INFO")
            url = self.browser_monitor._get_url_from_chrome_devtools()
            if url:
                self.log_message(f"✅ URL detected: {url}", "SUCCESS")
                try:
                    self.browser_monitor._handle_url_change(url)
                    base_url = self.browser_monitor.base_url or "Not set yet"
                    # Always show popup when URL is detected
                    self.root.after(0, lambda: messagebox.showinfo("URL Detected", 
                              f"URL successfully detected and updated!\n\n"
                              f"URL: {url}\n"
                              f"Base URL: {base_url}\n"
                              f"Module: {self.current_module}\n"
                              f"Page: {self.current_page}\n\n"
                              f"You can now start capturing test cases."))
                except Exception as e:
                    self.log_message(f"❌ Error handling URL change: {e}", "ERROR")
                    self.root.after(0, lambda: messagebox.showerror("Error", 
                                  f"Error updating URL:\n{e}\n\n"
                                  f"Detected URL: {url}\n\n"
                                  f"Please try using 'Detect URL from Browser' or 'Paste URL Manually' button."))
            else:
                if attempt < max_attempts:
                    # Retry after 2 more seconds
                    self.log_message(f"URL not detected yet. Retrying in 2 seconds... (attempt {attempt}/{max_attempts})", "INFO")
                    self.root.after(2000, lambda: self._try_detect_url_after_launch(attempt=attempt+1, max_attempts=max_attempts))
                else:
                    # Always show popup when URL is not detected after all attempts
                    self.log_message("URL not detected after multiple attempts.", "WARNING")
                    self.log_message("Please navigate to the target page in the browser, then:", "INFO")
                    self.log_message("1. Click 'Detect URL from Browser' button, OR", "INFO")
                    self.log_message("2. Copy URL from browser and use 'Paste URL Manually' button, OR", "INFO")
                    self.log_message("3. Use Manual Override section to enter URL manually", "INFO")
                    self.root.after(0, lambda: messagebox.showinfo("URL Not Detected", 
                                  "Could not automatically detect the URL from the launched browser.\n\n"
                                  "This is normal if:\n"
                                  "• Browser is still loading\n"
                                  "• Browser opened to a blank page\n"
                                  "• Remote debugging connection needs more time\n\n"
                                  "Please:\n"
                                  "1. Navigate to your target application in the browser\n"
                                  "2. Click 'Detect URL from Browser' or 'Paste URL Manually' button\n"
                                  "3. Or use Manual Override section"))
        except Exception as e:
            self.log_message(f"❌ Error in URL detection: {e}", "ERROR")
            self.root.after(0, lambda: messagebox.showerror("Error", 
                              f"Error detecting URL:\n{e}\n\n"
                              f"Please try:\n"
                              f"1. Click 'Detect URL from Browser' button\n"
                              f"2. Or use 'Paste URL Manually' button\n"
                              f"3. Or use Manual Override section"))
    
    def add_navigation_action(self, action):
        """Add navigation action to list"""
        step_number = len(self.actions_listbox.get(0, tk.END)) + 1
        self.actions_listbox.insert(tk.END, f"{step_number}. {action}")
        self.actions_listbox.see(tk.END)
        count = len(self.actions_listbox.get(0, tk.END))
        self.action_count_label.config(text=f"Actions captured: {count}")
    
    def start_monitoring(self):
        """Start automatic action monitoring"""
        if not self.browser_monitor.monitoring:
            self.log_message("Browser monitoring not started. Please start it first from Setup tab", "WARNING")
            messagebox.showwarning("Warning", "Browser monitoring not started! Please start it first from Setup tab.")
            return
        
        if not PYNPUT_AVAILABLE:
            self.log_message("pynput library not installed", "ERROR")
            messagebox.showerror("Error", 
                "pynput library not installed!\n\n"
                "Please install it using:\n"
                "pip install pynput")
            return
        
        self.log_message("Starting action monitoring...", "INFO")
        if self.monitor.start_monitoring():
            self.monitoring_active = True
            self.monitor_status_label.config(text="Status: Monitoring ON", foreground="green")
            self.start_monitor_btn.config(state=tk.DISABLED)
            self.stop_monitor_btn.config(state=tk.NORMAL)
            self.auto_save_enabled = self.auto_save_var.get()
            self.status_label.config(text="Auto-capture ACTIVE - Actions are being captured automatically!")
            self.log_message("Action monitoring started successfully", "SUCCESS")
            self.log_message("Mouse and keyboard listeners active", "INFO")
            self.log_message("✅ All actions (clicks, typing, scrolling, dropdowns, menus) will be captured", "SUCCESS")
            if self.current_url:
                self.log_message(f"Monitoring actions on: {self.current_url}", "INFO")
                if self.manual_url_set:
                    self.log_message("✓ URL was set manually - ALL clicks will be captured!", "SUCCESS")
                else:
                    self.log_message("⚠ URL auto-detected. If clicks aren't captured, set URL manually", "WARNING")
            else:
                self.log_message("⚠ URL not set. Set URL manually in 'Manual Override' section to capture clicks", "WARNING")
                self.log_message("💡 Tip: Copy URL from browser and paste in Manual Override, then click 'Set URL'", "INFO")
            
            # Important: Log that monitoring will continue after login/navigation
            self.log_message("=" * 60, "INFO")
            self.log_message("✅ Monitoring will continue after login and page navigation", "SUCCESS")
            self.log_message("✅ All actions will be logged in Activity Log below", "SUCCESS")
            self.log_message("=" * 60, "INFO")
        else:
            self.log_message("Failed to start action monitoring", "ERROR")
            messagebox.showerror("Error", "Failed to start monitoring!")
    
    def stop_monitoring(self):
        """Stop automatic action monitoring"""
        self.log_message("Stopping action monitoring...", "INFO")
        self.monitor.stop_monitoring()
        self.monitoring_active = False
        self.monitor_status_label.config(text="Status: Monitoring OFF", foreground="red")
        self.start_monitor_btn.config(state=tk.NORMAL)
        self.stop_monitor_btn.config(state=tk.DISABLED)
        self.status_label.config(text="Auto-capture STOPPED")
        self.log_message("Action monitoring stopped", "INFO")
    
    def on_action_captured(self, action):
        """Callback when an action is automatically captured"""
        # Always log actions, even if monitoring seems inactive (might be a timing issue)
        if not self.monitoring_active:
            self.log_message("⚠️ Action received but monitoring appears inactive - checking status...", "WARNING")
            # Don't return - still log and capture the action
        
        # Log the action capture with more detail
        action_lower = action.lower()
        if "dropdown" in action_lower or "menu" in action_lower:
            self.log_message(f"📋 DROPDOWN/MENU ACTION: {action}", "ACTION")
        elif "switch" in action_lower and "tab" in action_lower:
            self.log_message(f"🔄 TAB SWITCH ACTION: {action}", "ACTION")
        elif "click" in action_lower:
            self.log_message(f"🖱️ CLICK ACTION: {action}", "ACTION")
        elif "text" in action_lower or "typing" in action_lower:
            self.log_message(f"⌨️ TEXT INPUT: {action}", "ACTION")
        elif "navigate" in action_lower or "window" in action_lower:
            self.log_message(f"🌐 NAVIGATION: {action}", "ACTION")
        else:
            self.log_message(f"📝 ACTION CAPTURED: {action}", "ACTION")
        
        # Add to listbox
        step_number = len(self.actions_listbox.get(0, tk.END)) + 1
        self.actions_listbox.insert(tk.END, f"{step_number}. {action}")
        self.actions_listbox.see(tk.END)  # Scroll to bottom
        
        # Update count
        count = len(self.actions_listbox.get(0, tk.END))
        self.action_count_label.config(text=f"Actions captured: {count}")
        
        # Auto-save if enabled and threshold reached
        if self.auto_save_enabled and count > 0 and count % self.auto_save_interval == 0:
            self.log_message(f"Auto-save threshold reached ({count} actions). Saving test case...", "INFO")
            self.root.after(1000, self.auto_save_test_case)  # Save after 1 second delay
    
    def add_manual_action(self):
        """Add a manual action"""
        action = self.manual_action_entry.get().strip()
        if action:
            step_number = len(self.actions_listbox.get(0, tk.END)) + 1
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.actions_listbox.insert(tk.END, f"{step_number}. [{timestamp}] {action}")
            self.manual_action_entry.delete(0, tk.END)
            count = len(self.actions_listbox.get(0, tk.END))
            self.action_count_label.config(text=f"Actions captured: {count}")
            self.actions_listbox.see(tk.END)  # Scroll to bottom
            
            # Log manual action
            action_lower = action.lower()
            if "dropdown" in action_lower or "menu" in action_lower:
                self.log_message(f"📋 Manual dropdown/menu action added: {action}", "ACTION")
            elif "tab" in action_lower:
                self.log_message(f"🔄 Manual tab action added: {action}", "ACTION")
            else:
                self.log_message(f"📝 Manual action added: {action}", "ACTION")
    
    def add_template_action(self, template):
        """Add a template action"""
        step_number = len(self.actions_listbox.get(0, tk.END)) + 1
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.actions_listbox.insert(tk.END, f"{step_number}. [{timestamp}] {template}")
        count = len(self.actions_listbox.get(0, tk.END))
        self.action_count_label.config(text=f"Actions captured: {count}")
        self.actions_listbox.see(tk.END)  # Scroll to bottom
        
        # Enhanced logging for template actions
        if "dropdown" in template.lower() or "menu" in template.lower():
            self.log_message(f"📋 DROPDOWN/MENU ACTION: {template}", "ACTION")
        elif "tab" in template.lower():
            self.log_message(f"🔄 TAB ACTION: {template}", "ACTION")
        else:
            self.log_message(f"📝 Template action added: {template}", "ACTION")
        
        # Focus on manual action entry so user can edit if needed
        self.manual_action_entry.focus()
    
    def update_current_tab(self):
        """Update current tab name"""
        tab_name = self.current_tab_var.get().strip()
        if tab_name:
            old_tab = self.current_tab
            self.current_tab = tab_name
            if old_tab and old_tab != tab_name:
                self.log_message(f"🔄 Tab updated: {old_tab} → {tab_name}", "INFO")
            else:
                self.log_message(f"📑 Current tab set to: {tab_name}", "INFO")
            self.status_label.config(text=f"Current tab: {tab_name}")
            # Update session info to show tab
            self.update_session_info()
    
    def capture_tab_switch(self):
        """Capture a tab switch action"""
        new_tab = self.current_tab_var.get().strip()
        
        if not new_tab:
            # Show dialog to enter tab name
            response = messagebox.askyesno(
                "Tab Name Required",
                "Please enter the tab name in the 'Current Tab' field first.\n\n"
                "For example: 'Accounts', 'Users', 'Notification'\n\n"
                "Would you like to enter it now?"
            )
            if response:
                self.current_tab_entry.focus()
                return
            else:
                return
        
        # Log tab switch prominently with multiple log entries for visibility
        if self.previous_tab and self.previous_tab != new_tab:
            action = f"Switched from '{self.previous_tab}' tab to '{new_tab}' tab"
            # Multiple log entries to make it very visible
            self.log_message("=" * 60, "INFO")
            self.log_message(f"🔄 TAB SWITCH DETECTED: {self.previous_tab} → {new_tab}", "ACTION")
            self.log_message(f"Tab switch: {self.previous_tab} → {new_tab}", "SUCCESS")
            self.log_message(f"This tab switch will be included in the test case", "INFO")
            self.log_message("=" * 60, "INFO")
        else:
            action = f"Switched to '{new_tab}' tab"
            self.log_message("=" * 60, "INFO")
            self.log_message(f"🔄 TAB NAVIGATION: {new_tab}", "ACTION")
            self.log_message(f"Tab navigation to: {new_tab}", "SUCCESS")
            self.log_message(f"This tab navigation will be included in the test case", "INFO")
            self.log_message("=" * 60, "INFO")
        
        # Add to actions list
        step_number = len(self.actions_listbox.get(0, tk.END)) + 1
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.actions_listbox.insert(tk.END, f"{step_number}. [{timestamp}] {action}")
        count = len(self.actions_listbox.get(0, tk.END))
        self.action_count_label.config(text=f"Actions captured: {count}")
        self.actions_listbox.see(tk.END)
        
        # Update tab tracking
        self.previous_tab = self.current_tab
        self.current_tab = new_tab
        
        # Update page name to include tab if URL is same
        if self.current_page and new_tab:
            # Don't duplicate tab name if already in page name
            if new_tab not in self.current_page:
                self.current_page = f"{self.current_page} - {new_tab}"
            self.update_session_info()
        
        self.log_message(f"✅ Tab switch successfully captured and will be included in test case", "SUCCESS")
    
    def remove_action(self):
        """Remove selected action"""
        selection = self.actions_listbox.curselection()
        if selection:
            self.actions_listbox.delete(selection[0])
            # Renumber actions
            items = list(self.actions_listbox.get(0, tk.END))
            self.actions_listbox.delete(0, tk.END)
            for i, item in enumerate(items, 1):
                # Remove old number and add new one
                action_text = item.split('. ', 1)[1] if '. ' in item else item
                self.actions_listbox.insert(tk.END, f"{i}. {action_text}")
            count = len(self.actions_listbox.get(0, tk.END))
            self.action_count_label.config(text=f"Actions captured: {count}")
    
    def log_message(self, message, level="INFO"):
        """Add a log message to the log panel"""
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]  # Include milliseconds
        log_entry = f"[{timestamp}] [{level}] {message}\n"
        
        # Add to log storage
        self.log_messages.append((timestamp, level, message))
        if len(self.log_messages) > self.max_log_lines:
            self.log_messages.pop(0)
        
        # Update UI (thread-safe) - check if log_text exists
        try:
            if hasattr(self, 'log_text') and self.log_text:
                self.root.after(0, lambda: self._update_log_ui(log_entry, level))
        except:
            pass  # Log panel not ready yet
    
    def _update_log_ui(self, log_entry, level):
        """Update log UI (called from main thread)"""
        try:
            self.log_text.insert(tk.END, log_entry)
            # Apply color tag based on level
            start_pos = self.log_text.index(tk.END + "-1c linestart")
            end_pos = self.log_text.index(tk.END + "-1c lineend")
            self.log_text.tag_add(level, start_pos, end_pos)
            self.log_text.see(tk.END)  # Auto-scroll to bottom
            
            # Limit log size in UI
            lines = int(self.log_text.index('end-1c').split('.')[0])
            if lines > self.max_log_lines:
                self.log_text.delete('1.0', f'{lines - self.max_log_lines}.0')
        except:
            pass
    
    def clear_logs(self):
        """Clear the log panel"""
        self.log_text.delete(1.0, tk.END)
        self.log_messages.clear()
        self.log_message("Logs cleared", "INFO")
    
    def clear_actions(self):
        """Clear all captured actions"""
        self.actions_listbox.delete(0, tk.END)
        self.action_count_label.config(text="Actions captured: 0")
        self.expected_result_text.delete(1.0, tk.END)
        self.actual_result_text.delete(1.0, tk.END)
        self.status_var.set("Not Executed")
        self.log_message("Actions cleared", "INFO")
    
    def auto_save_test_case(self):
        """Automatically save test case when threshold is reached"""
        if len(self.actions_listbox.get(0, tk.END)) >= self.auto_save_interval:
            # Use default values for auto-save
            expected_result = self.expected_result_text.get(1.0, tk.END).strip() or self._generate_expected_result()
            actual_result = self.actual_result_text.get(1.0, tk.END).strip() or "Captured automatically"
            status = self.status_var.get()
            
            self.save_test_case_internal(expected_result, actual_result, status, silent=True)
            
            # Clear actions for next batch
            self.clear_actions()
    
    def _generate_expected_result(self):
        """Generate expected result based on navigation and actions"""
        actions_text = ' '.join(self.actions_listbox.get(0, tk.END)).lower()
        
        if self.current_module and self.current_page:
            # Check if tab switching is involved
            if "switch" in actions_text and "tab" in actions_text:
                if self.current_tab:
                    return f"User should successfully switch to '{self.current_tab}' tab and see the {self.current_tab} content displayed correctly"
                else:
                    return f"User should successfully switch tabs and see the updated content"
            
            # Check if dropdown/menu actions are involved
            if "dropdown" in actions_text or "menu" in actions_text:
                if "open" in actions_text:
                    return f"Dropdown/menu should open successfully and display available options"
                elif "select" in actions_text:
                    return f"Option should be selected from dropdown/menu and applied successfully"
                else:
                    return f"Dropdown/menu interaction should work correctly"
            
            if "login" in self.current_module.lower():
                return f"User should be successfully logged in and redirected to {self.current_page}"
            elif "navigate" in actions_text:
                return f"User should be navigated to {self.current_page} page"
            else:
                page_desc = f"{self.current_page} - {self.current_tab}" if self.current_tab else self.current_page
                return f"Action should be completed successfully on {page_desc} page"
        return "Action should be completed successfully"
    
    def set_manual_url(self):
        """Set URL manually"""
        url = self.manual_url_var.get().strip()
        if url:
            self.log_message(f"Setting URL manually: {url}", "INFO")
            if not url.startswith('http'):
                url = 'https://' + url
                self.log_message(f"Added https:// prefix. Final URL: {url}", "INFO")
            self.current_url = url
            self.browser_monitor.current_url = url  # Update browser monitor
            self.manual_url_set = True  # Mark as manually set
            self.monitor.manual_url_set = True  # Pass to action monitor
            
            # Try to identify module and page from URL (works for any URL now)
            # Base URL will be set automatically from the first URL
            module, page = self.browser_monitor._identify_module_and_page(url)
            self.current_module = module
            self.current_page = page
            self.browser_monitor.current_module = module
            self.browser_monitor.current_page = page
            self.manual_module_var.set(module)
            self.manual_page_var.set(page)
            self.update_url_info(url, module, page)
            self.status_label.config(text=f"✓ URL set: {url} | Module: {module} | Page: {page} | Ready to capture!")
            # Show success but don't block with messagebox - just update status
            base_url = self.browser_monitor.base_url or "Will be set from URL"
            self.url_status_label.config(text=f"✓ URL set successfully! Module: {module} | Page: {page} | You can now start capturing.", 
                                       foreground="green")
            # Update base URL label
            if hasattr(self, 'base_url_label'):
                self.base_url_label.config(text=base_url, foreground="blue", font=("Arial", 10, "bold"))
            self.log_message(f"URL set successfully. Module: {module}, Page: {page}", "SUCCESS")
            self.log_message(f"Base URL set to: {base_url}", "INFO")
            self.log_message("Actions will now be captured when you click in the browser!", "SUCCESS")
            if self.monitoring_active:
                self.log_message("Action monitoring is active. Actions will now be captured!", "SUCCESS")
                self.log_message("URL set despite not matching base URL (user confirmed)", "WARNING")
    
    def set_manual_module(self):
        """Set module manually"""
        module = self.manual_module_var.get().strip()
        if module:
            self.current_module = module
            self.update_session_info()
            self.status_label.config(text=f"Module set manually: {module}")
    
    def set_manual_page(self):
        """Set page manually"""
        page = self.manual_page_var.get().strip()
        if page:
            self.current_page = page
            self.update_session_info()
            self.status_label.config(text=f"Page set manually: {page}")
    
    def update_session_info(self):
        """Update session info label"""
        module = self.current_module or "Not set"
        page = self.current_page or "Not set"
        tab_info = f" | Tab: {self.current_tab}" if self.current_tab else ""
        self.session_info_label.config(text=f"Module: {module} | Page: {page}{tab_info}")
    
    def save_test_case(self):
        """Save current test case to Excel"""
        # Check if we have module - if not, try to use manual override or show helpful error
        if not self.current_module:
            # Check if manual module is set
            manual_module = self.manual_module_var.get().strip()
            if manual_module:
                self.current_module = manual_module
            else:
                # Show dialog to set module manually
                response = messagebox.askyesno(
                    "Module Not Detected",
                    "No module detected automatically.\n\n"
                    "Possible reasons:\n"
                    "1. Browser not started with --remote-debugging-port=9222\n"
                    "2. Not on the target URL\n"
                    "3. Browser monitoring not connected\n\n"
                    "Would you like to set the module manually?\n\n"
                    "Click 'Yes' to set manually, or 'No' to cancel."
                )
                if response:
                    # Focus on manual module field
                    self.manual_module_var.set("General")
                    self.current_module = "General"
                    if not self.current_page:
                        manual_page = self.manual_page_var.get().strip() or "Home"
                        self.current_page = manual_page
                        self.manual_page_var.set(manual_page)
                    if not self.current_url:
                        # Don't set a hardcoded URL - user should set it manually
                        self.log_message("Please set the URL using 'Detect URL from Browser' or 'Paste URL Manually' button", "INFO")
                    self.update_session_info()
                    messagebox.showinfo("Module Set", 
                                      f"Module set to: {self.current_module}\n"
                                      f"Page set to: {self.current_page}\n"
                                      f"You can change these in the 'Manual Override' section if needed.")
                else:
                    return
        
        # Get test steps
        test_steps = '\n'.join(self.actions_listbox.get(0, tk.END))
        if not test_steps:
            # Allow saving with no actions, but show a warning
            response = messagebox.askyesno(
                "No Actions Captured",
                "No actions have been captured yet.\n\n"
                "You can:\n"
                "1. Add manual actions using the 'Add Manual Action' field\n"
                "2. Start monitoring and perform actions in your application\n"
                "3. Save with no actions (not recommended)\n\n"
                "Would you like to:\n"
                "- Click 'Yes' to add manual actions first\n"
                "- Click 'No' to save without actions (you can add them later)"
            )
            if response:
                # Focus on manual action entry
                self.manual_action_entry.focus()
                messagebox.showinfo("Add Manual Actions", 
                                  "Please add manual actions in the 'Add Manual Action' field above.\n"
                                  "You can add multiple actions, then click 'Save Test Case to Excel' again.")
                return
            else:
                # User chose to save without actions - use a default message
                test_steps = "No actions captured yet. Please add test steps manually."
        
        expected_result = self.expected_result_text.get(1.0, tk.END).strip() or self._generate_expected_result()
        actual_result = self.actual_result_text.get(1.0, tk.END).strip()
        status = self.status_var.get()
        
        self.save_test_case_internal(expected_result, actual_result, status, silent=False)
    
    def save_test_case_internal(self, expected_result, actual_result, status, silent=False):
        """Internal method to save test case"""
        # Get test steps
        test_steps = '\n'.join(self.actions_listbox.get(0, tk.END))
        
        # Get functionality
        functionality = self.functionality_text.get(1.0, tk.END).strip() or f"{self.current_module} - {self.current_page}"
        
        # Initialize module if needed
        if self.current_module not in self.test_cases_by_module:
            self.test_cases_by_module[self.current_module] = []
            self.test_case_counters[self.current_module] = 0
        
        # Generate test case ID
        module_short = self.current_module.upper().replace(' ', '_')[:20]
        self.test_case_counters[self.current_module] += 1
        test_id = f"TC_{module_short}_{self.test_case_counters[self.current_module]:03d}"
        
        self.log_message(f"Saving test case: {test_id}", "INFO")
        self.log_message(f"Module: {self.current_module}, Page: {self.current_page}", "INFO")
        if self.current_tab:
            self.log_message(f"📑 Tab: {self.current_tab}", "INFO")
        self.log_message(f"Test steps: {len(self.actions_listbox.get(0, tk.END))} actions", "INFO")
        
        # Check if test case includes tab switches
        test_steps_text = test_steps.lower()
        if "switch" in test_steps_text and "tab" in test_steps_text:
            self.log_message(f"✅ Test case includes tab switch actions", "SUCCESS")
        
        # Check if test case includes dropdown/menu actions
        if "dropdown" in test_steps_text or "menu" in test_steps_text:
            self.log_message(f"✅ Test case includes dropdown/menu actions", "SUCCESS")
        
        # Include tab information in test case if available
        page_name = self.current_page
        if self.current_tab and self.current_tab not in page_name:
            page_name = f"{page_name} - {self.current_tab}"
        
        # Create test case
        test_case = {
            "test_id": test_id,
            "test_name": f"Verify {functionality} on {page_name}",
            "description": f"Test {functionality} functionality on {page_name} page",
            "preconditions": f"User is on {page_name} page (URL: {self.current_url})",
            "test_steps": test_steps,
            "expected_result": expected_result or self._generate_expected_result(),
            "actual_result": actual_result or "",
            "status": status,
            "priority": "High" if status == "Pass" else "Medium",
            "module": self.current_module,
            "page": page_name,
            "url": self.current_url,
            "tab": self.current_tab if self.current_tab else "",
            "created_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # Add to test cases
        self.test_cases_by_module[self.current_module].append(test_case)
        
        # Save to Excel
        try:
            self.log_message("Exporting to Excel...", "INFO")
            self.export_to_excel()
            self.log_message(f"Test case saved successfully to Excel: {self.excel_file_path}", "SUCCESS")
            
            # Update status
            if not silent:
                self.status_label.config(
                    text=f"Test case {test_id} saved to {self.current_module} module! Total: {len(self.test_cases_by_module[self.current_module])} test cases in this module")
                messagebox.showinfo("Success", 
                                  f"Test case {test_id} saved to {self.current_module} module in {self.excel_file_path}!")
            else:
                self.status_label.config(
                    text=f"Auto-saved test case {test_id} to {self.current_module} module!")
                self.log_message(f"Auto-saved test case {test_id}", "SUCCESS")
        except Exception as e:
            self.log_message(f"Error saving test case: {str(e)}", "ERROR")
            if not silent:
                messagebox.showerror("Error", f"Failed to save test case: {str(e)}")
    
    def export_to_excel(self):
        """Export all test cases to Excel file organized by module"""
        try:
            # Try to load existing workbook
            try:
                wb = load_workbook(self.excel_file_path)
            except FileNotFoundError:
                wb = Workbook()
                wb.remove(wb.active)  # Remove default sheet
            
            # Define headers - new structure as per requirements
            headers = [
                "TC_ID", "TC_Module", "Prerequisite", "Execution_Steps",
                "Expected_Output", "Actual_Output", "Status", "Priority",
                "URL", "Created Date"
            ]
            
            # Style for header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Create/update sheets for each module
            for module, test_cases in self.test_cases_by_module.items():
                if not test_cases:
                    continue
                
                # Create or get sheet for module
                sheet_name = module[:31]  # Excel sheet name limit
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Clear existing data (keep header)
                    if ws.max_row > 1:
                        ws.delete_rows(2, ws.max_row)
                else:
                    ws = wb.create_sheet(title=sheet_name)
            
                # Write headers for this sheet
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
                # Write test cases for this module - new column structure
                for row_num, test_case in enumerate(test_cases, 2):
                    # Column 1: TC_ID
                    ws.cell(row=row_num, column=1, value=test_case.get("test_id", "")).border = border
                    # Column 2: TC_Module
                    ws.cell(row=row_num, column=2, value=test_case.get("module", "")).border = border
                    # Column 3: Prerequisite
                    ws.cell(row=row_num, column=3, value=test_case.get("preconditions", "")).border = border
                    # Column 4: Execution_Steps
                    ws.cell(row=row_num, column=4, value=test_case.get("test_steps", "")).border = border
                    # Column 5: Expected_Output
                    ws.cell(row=row_num, column=5, value=test_case.get("expected_result", "")).border = border
                    # Column 6: Actual_Output
                    ws.cell(row=row_num, column=6, value=test_case.get("actual_result", "")).border = border
                    # Column 7: Status
                    ws.cell(row=row_num, column=7, value=test_case.get("status", "")).border = border
                    # Column 8: Priority
                    ws.cell(row=row_num, column=8, value=test_case.get("priority", "")).border = border
                    # Column 9: URL
                    ws.cell(row=row_num, column=9, value=test_case.get("url", "")).border = border
                    # Column 10: Created Date
                    ws.cell(row=row_num, column=10, value=test_case.get("created_date", "")).border = border
                    
                    # Color code status (now in column 7)
                    status_cell = ws.cell(row=row_num, column=7)
                status = test_case.get("status", "")
                if status == "Pass":
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif status == "Fail":
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif status == "Blocked":
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
                # Auto-adjust column widths for this sheet
                for col_num, header in enumerate(headers, 1):
                    max_length = len(header)
                    for row_num in range(2, len(test_cases) + 2):
                        cell_value = ws.cell(row=row_num, column=col_num).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)
            
                # Enable text wrapping for columns with long text
                for row_num in range(2, len(test_cases) + 2):
                    for col_num in [3, 4, 5, 6]:  # Prerequisite, Execution_Steps, Expected_Output, Actual_Output
                        ws.cell(row=row_num, column=col_num).alignment = Alignment(
                            wrap_text=True, vertical="top"
                        )
            
            # Freeze header row
            ws.freeze_panes = "A2"
            
            # Save file
            wb.save(self.excel_file_path)
            
        except Exception as e:
            raise Exception(f"Failed to export to Excel: {str(e)}")


def main():
    root = tk.Tk()
    app = TestCaseCapture(root)
    root.mainloop()


if __name__ == "__main__":
    main()
